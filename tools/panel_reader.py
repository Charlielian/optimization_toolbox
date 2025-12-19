# -*- coding: utf-8 -*-
"""
优化百宝箱工具集 - 面板数据分析工具
提供网格评估、方案分析、数据导出等功能
"""

import os
import sys
import datetime
import time
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import streamlit as st
import uuid
import json
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import matplotlib.pyplot as plt
from matplotlib import font_manager
import numpy as np


class PanelReader:
    """面板数据分析工具"""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.logger = logging.getLogger(__name__)
        
    def render(self):
        """渲染面板数据分析界面"""
        st.title("📊 面板数据分析")
        st.caption("网格评估、方案分析、数据导出功能")
        
        # 功能导航
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📁 数据读取", "📊 网格评估", "📈 方案分析", "📤 数据导出", "🔍 数据查询"
        ])
        
        with tab1:
            self._render_data_reading()
            
        with tab2:
            self._render_grid_evaluation()
            
        with tab3:
            self._render_scheme_analysis()
            
        with tab4:
            self._render_data_export()
            
        with tab5:
            self._render_data_query()
    
    def _render_data_reading(self):
        """渲染数据读取页面"""
        st.subheader("📁 数据读取")
        
        # 创建导入类型选择
        import_type = st.radio(
            "选择导入类型",
            ["网格数据导入", "网格结果得分导入"],
            horizontal=True
        )
        
        if import_type == "网格数据导入":
            self._render_panel_data_import()
        else:
            self._render_grid_result_score_import()
    
    def _render_panel_data_import(self):
        """渲染面板数据导入（整合网格数据导入和超时方案清单导入）"""
        st.markdown("#### 上传网格数据文件")
        st.info("""
        **支持两种格式的CSV文件：**
        1. **total文件格式**：包含vcorder_code、business_key_、current_act_name等字段的传统格式
        2. **超时方案清单格式**：包含工单编号、优化单号、方案提交时间等字段的标准格式（推荐）
        
        系统会自动识别文件格式并进行相应的列名映射处理。
        """)
        
        uploaded_files = st.file_uploader(
            "选择CSV文件（支持total文件和超时方案清单文件）",
            type=['csv', 'xlsx', 'xls'],
            accept_multiple_files=True,
            help="支持同时上传多个CSV/Excel文件，系统会自动识别文件格式",
            key="panel_data_files"
        )
        
        if uploaded_files:
            st.success(f"已选择 {len(uploaded_files)} 个文件")
            
            # 显示文件列表
            with st.expander("查看文件列表"):
                for i, file in enumerate(uploaded_files, 1):
                    st.write(f"{i}. {file.name}")
            
            if st.button("开始读取数据", type="primary", key="import_panel_data"):
                self._process_uploaded_files(uploaded_files)
    
    def _render_timeout_scheme_import(self):
        """渲染超时方案清单导入"""
        st.markdown("#### 导入超时方案清单")
        st.info("支持超时方案清单数据导入，文件需包含'方案id'字段，将自动与panel_data表的scheme_id关联")
        
        timeout_files = st.file_uploader(
            "选择超时方案清单文件",
            type=['xlsx', 'xls', 'csv'],
            accept_multiple_files=True,
            help="支持同时上传多个超时方案清单文件",
            key="timeout_scheme_files"
        )
        
        if timeout_files:
            st.success(f"已选择 {len(timeout_files)} 个文件")
            
            # 显示文件列表
            with st.expander("查看文件列表"):
                for i, file in enumerate(timeout_files, 1):
                    st.write(f"{i}. {file.name}")
            
            if st.button("开始导入超时方案清单", type="primary", key="import_timeout_scheme"):
                self._batch_import_timeout_scheme_data(timeout_files)
    
    def _render_grid_result_score_import(self):
        """渲染网格结果得分导入"""
        st.markdown("#### 导入网格结果得分")
        st.info("支持CSV格式的网格结果得分文件，包含时间周期、网格ID、各项得分指标等信息")
        
        score_files = st.file_uploader(
            "选择网格结果得分CSV文件",
            type=['csv'],
            accept_multiple_files=True,
            help="支持同时上传多个CSV文件，文件需包含'时间'和'微网格'字段",
            key="grid_result_score_files"
        )
        
        if score_files:
            st.success(f"已选择 {len(score_files)} 个文件")
            
            # 显示文件列表
            with st.expander("查看文件列表"):
                for i, file in enumerate(score_files, 1):
                    st.write(f"{i}. {file.name}")
            
            if st.button("开始导入网格结果得分", type="primary", key="import_grid_result_score"):
                self._batch_import_grid_result_score_data(score_files)
    
    def _read_excel_or_csv(self, uploaded_file):
        """读取Excel或CSV文件"""
        import io
        name = uploaded_file.name.lower()
        
        if name.endswith(('.xlsx', '.xls')):
            try:
                # 重置文件指针
                uploaded_file.seek(0)
                
                # 读取Excel文件
                df = pd.read_excel(
                    uploaded_file, 
                    engine='openpyxl',
                    na_values=['', 'NA', 'N/A', 'null'],
                    keep_default_na=False
                )
                return df
            except Exception as e:
                error_msg = f"读取Excel文件 '{uploaded_file.name}' 失败: {str(e)}"
                self.logger.error(error_msg)
                raise Exception(error_msg)
                
        elif name.endswith('.csv'):
            try:
                uploaded_file.seek(0)
                content = uploaded_file.read()
                df = pd.read_csv(io.BytesIO(content), encoding='utf-8')
                return df
            except Exception as e:
                # 尝试其他编码
                try:
                    uploaded_file.seek(0)
                    content = uploaded_file.read()
                    df = pd.read_csv(io.BytesIO(content), encoding='gbk')
                    return df
                except Exception as e2:
                    raise Exception(f"读取CSV文件失败: {str(e)}")
        else:
            raise Exception(f"不支持的文件格式: {name}")
    
    def _batch_import_timeout_scheme_data(self, files):
        """批量导入超时方案清单数据"""
        try:
            total_files = len(files)
            total_ok = 0
            total_err = 0
            
            st.write("🚀 **开始超时方案清单导入过程**")
            st.write(f"📁 待处理文件数量: {total_files}")
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for i, file in enumerate(files):
                file_start_time = time.time()
                status_text.text(f"正在处理文件 {i + 1}/{total_files}: {file.name}")
                
                try:
                    st.write(f"📖 正在读取文件: {file.name}...")
                    
                    # 读取文件
                    df = self._read_excel_or_csv(file)
                    st.success(f"✅ 文件读取成功！共 {len(df):,} 行数据")
                    
                    # 显示文件信息
                    st.write(f"📄 **文件 {i + 1}/{total_files}: {file.name}**")
                    st.write(f"  • 数据行数: {len(df):,} 行")
                    st.write(f"  • 列数: {len(df.columns)} 列")
                    
                    # 检查必要字段：方案id
                    if '方案id' not in df.columns:
                        st.error(f"❌ 文件缺少必要字段: '方案id'")
                        total_err += 1
                        continue
                    
                    # 处理超时方案清单数据
                    st.write("🔄 **开始处理超时方案清单数据**")
                    records = []
                    error_rows = 0
                    
                    for idx, row in df.iterrows():
                        try:
                            # 必须字段：方案id
                            scheme_id = str(row['方案id']).strip()
                            
                            if not scheme_id or scheme_id == 'nan' or scheme_id == '':
                                error_rows += 1
                                continue
                            
                            # 过滤掉24年完成的方案ID（不参与过程分计算）
                            if self.db_manager.is_scheme_excluded(scheme_id):
                                continue  # 跳过这个方案，不导入
                            
                            # 处理其他字段
                            order_number = str(row.get('工单编号', '')).strip()
                            optimize_number = str(row.get('优化单号', '')).strip()
                            process_status = str(row.get('流程状态', '')).strip()
                            start_time = str(row.get('开始时间', '')).strip()
                            grid_code = str(row.get('微网格编号', '')).strip()
                            grid_name = str(row.get('微网格名称', '')).strip()
                            label = str(row.get('标签', '')).strip()
                            city = str(row.get('地市', '')).strip()
                            district = str(row.get('区县', '')).strip()
                            reason_category = str(row.get('原因分类', '')).strip()
                            root_cause = str(row.get('根本原因', '')).strip()
                            scheme_category = str(row.get('方案分类', '')).strip()
                            measures = str(row.get('措施', '')).strip()
                            scheme_type = str(row.get('方案类型', '')).strip()
                            cell_name = str(row.get('小区名称', '')).strip()
                            adjust_param = str(row.get('调整参数', '')).strip()
                            adjust_before_value = str(row.get('调整前值', '')).strip()
                            target_value = str(row.get('目标值', '')).strip()
                            sub_order_number = str(row.get('子工单号', '')).strip()
                            sub_order_status = str(row.get('子工单状态', '')).strip()
                            implement_results = str(row.get('实施结果', '')).strip()
                            exclude_status = str(row.get('剔除/线下完成', '')).strip()
                            update_label = str(row.get('更新标签', '')).strip()
                            scheme_submit_time = str(row.get('方案提交时间', '')).strip()
                            scheme_complete_time = str(row.get('方案完成时间', '')).strip()
                            scheme_status = str(row.get('方案状态', '')).strip()
                            
                            # 处理数值字段
                            scheme_execution_time = None
                            scheme_standard_time = None
                            try:
                                if row.get('方案执行耗时') not in [None, '', 'nan']:
                                    scheme_execution_time = float(row['方案执行耗时'])
                            except (ValueError, TypeError):
                                pass
                            
                            try:
                                if row.get('方案标准时长') not in [None, '', 'nan']:
                                    scheme_standard_time = float(row['方案标准时长'])
                            except (ValueError, TypeError):
                                pass
                            
                            records.append((
                                scheme_id, order_number, optimize_number, process_status, start_time,
                                grid_code, grid_name, label, city, district, reason_category, root_cause,
                                scheme_category, measures, scheme_type, cell_name, adjust_param,
                                adjust_before_value, target_value, sub_order_number, sub_order_status,
                                implement_results, exclude_status, update_label, scheme_submit_time,
                                scheme_complete_time, scheme_execution_time, scheme_standard_time, scheme_status
                            ))
                            
                        except Exception as e:
                            error_rows += 1
                            continue
                    
                    if not records:
                        st.warning("⚠️ 没有有效的数据可以导入")
                        continue
                    
                    st.write(f"✅ 数据处理完成: {len(records):,} 条有效记录，{error_rows} 条错误行")
                    
                    # 显示前3条记录样本
                    if records:
                        with st.expander("查看数据样本（前3条）"):
                            for idx, record in enumerate(records[:3], 1):
                                st.write(f"第{idx}条: 方案id={record[0]}, 工单编号={record[1]}, 地市={record[8]}, 方案状态={record[27]}")
                    
                    # 批量插入数据库
                    st.write("💾 **开始批量数据库插入**")
                    insert_sql = """
                    INSERT OR REPLACE INTO timeout_scheme_list 
                    (scheme_id, order_number, optimize_number, process_status, start_time,
                     grid_code, grid_name, label, city, district, reason_category, root_cause,
                     scheme_category, measures, scheme_type, cell_name, adjust_param,
                     adjust_before_value, target_value, sub_order_number, sub_order_status,
                     implement_results, exclude_status, update_label, scheme_submit_time,
                     scheme_complete_time, scheme_execution_time, scheme_standard_time, scheme_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                    
                    success = self.db_manager.execute_many(insert_sql, records)
                    if success:
                        total_ok += len(records)
                        st.success(f"✅ 文件 {file.name} 导入成功！共 {len(records):,} 条记录")
                        
                        # 显示关联信息
                        st.info(f"ℹ️ 数据已导入，可以通过 scheme_id 与 panel_data 表关联查询")
                    else:
                        st.error(f"❌ 文件 {file.name} 数据库插入失败")
                        total_err += 1
                    
                except Exception as e:
                    error_msg = f"文件 {file.name} 处理失败: {str(e)}"
                    st.error(f"❌ {error_msg}")
                    self.logger.error(error_msg)
                    total_err += 1
                    continue
                
                # 更新进度
                progress = (i + 1) / total_files
                progress_bar.progress(progress)
            
            st.success(f"🎉 超时方案清单导入完成！成功: {total_ok:,} 条记录，失败: {total_err} 个文件")
            
        except Exception as e:
            st.error(f"超时方案清单导入过程失败: {str(e)}")
            self.logger.error(f"超时方案清单导入过程失败: {str(e)}")
    
    def _batch_import_grid_result_score_data(self, files):
        """批量导入网格结果得分数据"""
        try:
            total_files = len(files)
            total_ok = 0
            total_err = 0
            total_duplicate = 0
            
            st.write("🚀 **开始网格结果得分导入过程**")
            st.write(f"📁 待处理文件数量: {total_files}")
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for i, file in enumerate(files):
                file_start_time = time.time()
                status_text.text(f"正在处理文件 {i + 1}/{total_files}: {file.name}")
                
                try:
                    st.write(f"📖 正在读取文件: {file.name}...")
                    
                    # 读取文件
                    df = self._read_excel_or_csv(file)
                    st.success(f"✅ 文件读取成功！共 {len(df):,} 行数据")
                    
                    # 显示文件信息
                    st.write(f"📄 **文件 {i + 1}/{total_files}: {file.name}**")
                    st.write(f"  • 数据行数: {len(df):,} 行")
                    st.write(f"  • 列数: {len(df.columns)} 列")
                    
                    # 检查必要字段：时间、微网格
                    required_fields = ['时间', '微网格']
                    missing_fields = [f for f in required_fields if f not in df.columns]
                    if missing_fields:
                        st.error(f"❌ 文件缺少必要字段: {', '.join(missing_fields)}")
                        total_err += 1
                        continue
                    
                    # 处理网格结果得分数据
                    st.write("🔄 **开始处理网格结果得分数据**")
                    records = []
                    error_rows = 0
                    
                    for idx, row in df.iterrows():
                        try:
                            # 必须字段：时间、微网格
                            time_period = str(row['时间']).strip()
                            grid_code = str(row['微网格']).strip()
                            
                            if not time_period or time_period == 'nan' or time_period == '':
                                error_rows += 1
                                if error_rows <= 5:  # 记录前5个错误详情
                                    self.logger.warning(
                                        f"文件 {file.name} 第 {idx+2} 行跳过: 时间字段为空"
                                        f" (时间={repr(time_period)}, 微网格={repr(grid_code)})"
                                    )
                                continue
                            if not grid_code or grid_code == 'nan' or grid_code == '':
                                error_rows += 1
                                if error_rows <= 5:  # 记录前5个错误详情
                                    self.logger.warning(
                                        f"文件 {file.name} 第 {idx+2} 行跳过: 微网格字段为空"
                                        f" (时间={repr(time_period)}, 微网格={repr(grid_code)})"
                                    )
                                continue
                            
                            # 提取关键字段
                            province = str(row.get('省', '')).strip() if pd.notna(row.get('省')) else ''
                            grid_name = str(row.get('中文名', '')).strip() if pd.notna(row.get('中文名')) else ''
                            scene_detail = str(row.get('场景-细项', '')).strip() if pd.notna(row.get('场景-细项')) else ''
                            scene_merged = str(row.get('场景合并', '')).strip() if pd.notna(row.get('场景合并')) else ''
                            city = str(row.get('地市', '')).strip() if pd.notna(row.get('地市')) else ''
                            city_district = str(row.get('市公司-县区公司（唯一标识）', '')).strip() if pd.notna(row.get('市公司-县区公司（唯一标识）')) else ''
                            city_company = str(row.get('地市公司', '')).strip() if pd.notna(row.get('地市公司')) else ''
                            scene_area_attribute = str(row.get('场景区域属性', '')).strip() if pd.notna(row.get('场景区域属性')) else ''
                            group_name = str(row.get('分组', '')).strip() if pd.notna(row.get('分组')) else ''
                            supervise_label_2025 = str(row.get('2025年督办标签', '')).strip() if pd.notna(row.get('2025年督办标签')) else ''
                            
                            # 最终得分
                            final_score = None
                            if pd.notna(row.get('最终得分')):
                                try:
                                    final_score = float(row['最终得分'])
                                except (ValueError, TypeError):
                                    final_score = None
                            
                            # 投诉单量
                            complaint_count = None
                            if pd.notna(row.get('投诉单量')):
                                try:
                                    complaint_count = int(float(row['投诉单量']))
                                except (ValueError, TypeError):
                                    complaint_count = None
                            
                            # 日均RRC最大用户数
                            daily_max_rrc_users = None
                            if pd.notna(row.get('日均RRC最大用户数')):
                                try:
                                    daily_max_rrc_users = int(float(row['日均RRC最大用户数']))
                                except (ValueError, TypeError):
                                    daily_max_rrc_users = None
                            
                            # 将除关键字段外的所有数据存储为JSON
                            grid_result_data = {}
                            exclude_fields = ['时间', '微网格', '省', '中文名', '场景-细项', '场景合并', 
                                            '地市', '市公司-县区公司（唯一标识）', '地市公司', '场景区域属性', 
                                            '分组', '2025年督办标签', '最终得分', '投诉单量', '日均RRC最大用户数']
                            
                            for col in df.columns:
                                if col not in exclude_fields:
                                    value = row.get(col)
                                    if pd.notna(value):
                                        # 尝试转换为数字类型
                                        try:
                                            if isinstance(value, (int, float)):
                                                grid_result_data[col] = value
                                            else:
                                                value_str = str(value).strip()
                                                if value_str:
                                                    # 尝试转换为数字
                                                    try:
                                                        if '.' in value_str:
                                                            grid_result_data[col] = float(value_str)
                                                        else:
                                                            grid_result_data[col] = int(float(value_str))
                                                    except ValueError:
                                                        grid_result_data[col] = value_str
                                        except Exception:
                                            grid_result_data[col] = str(value) if pd.notna(value) else None
                                    else:
                                        grid_result_data[col] = None
                            
                            grid_result_data_json = json.dumps(grid_result_data, ensure_ascii=False)
                            
                            # 构建记录
                            record = {
                                'time_period': time_period,
                                'grid_code': grid_code,
                                'province': province if province else None,
                                'grid_name': grid_name if grid_name else None,
                                'scene_detail': scene_detail if scene_detail else None,
                                'scene_merged': scene_merged if scene_merged else None,
                                'city': city if city else None,
                                'city_district': city_district if city_district else None,
                                'city_company': city_company if city_company else None,
                                'scene_area_attribute': scene_area_attribute if scene_area_attribute else None,
                                'group_name': group_name if group_name else None,
                                'supervise_label_2025': supervise_label_2025 if supervise_label_2025 else None,
                                'final_score': final_score,
                                'complaint_count': complaint_count,
                                'daily_max_rrc_users': daily_max_rrc_users,
                                'grid_result_data_json': grid_result_data_json
                            }
                            
                            records.append(record)
                            
                        except Exception as e:
                            error_rows += 1
                            if error_rows <= 5:  # 只记录前5个错误
                                self.logger.warning(f"文件 {file.name} 第 {idx+1} 行处理失败: {str(e)}")
                            continue
                    
                    if error_rows > 0:
                        st.warning(
                            f"⚠️ 文件 {file.name} 有 {error_rows} 行数据处理失败"
                            f"（这些行的时间或微网格字段为空，已自动跳过）"
                        )
                    
                    if records:
                        st.write(f"💾 正在导入 {len(records):,} 条记录到数据库...")
                        
                        # 导入数据库
                        success_count, duplicate_count = self.db_manager.insert_grid_result_scores(records)
                        
                        if success_count > 0:
                            total_ok += success_count
                            total_duplicate += duplicate_count
                            st.success(f"✅ 文件 {file.name} 导入成功！新增: {success_count:,} 条记录，更新: {duplicate_count:,} 条记录")
                        else:
                            st.error(f"❌ 文件 {file.name} 数据库插入失败")
                            total_err += 1
                    else:
                        st.warning(f"⚠️ 文件 {file.name} 没有有效数据")
                        total_err += 1
                    
                except Exception as e:
                    error_msg = f"文件 {file.name} 处理失败: {str(e)}"
                    st.error(f"❌ {error_msg}")
                    self.logger.error(error_msg)
                    total_err += 1
                    continue
                
                # 更新进度
                progress = (i + 1) / total_files
                progress_bar.progress(progress)
            
            st.success(f"🎉 网格结果得分导入完成！成功: {total_ok:,} 条记录，更新: {total_duplicate:,} 条记录，失败: {total_err} 个文件")
            
        except Exception as e:
            st.error(f"网格结果得分导入过程失败: {str(e)}")
            self.logger.error(f"网格结果得分导入过程失败: {str(e)}")
    
    def _render_grid_evaluation(self):
        """渲染网格评估页面"""
        st.subheader("📊 网格评估")
        
        # 功能选择
        function_type = st.radio(
            "选择功能",
            ["网格评估", "网格分析"],
            horizontal=True
        )
        
        if function_type == "网格分析":
            self._render_grid_analysis()
            return
        
        # 评估参数设置
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 评估参数")
            kc_score = st.slider("勘察得分权重", 0.0, 0.1, 0.05, 0.01, help="勘察得分在总评分中的权重")
            gj_weight = st.slider("规建类权重", 0.0, 0.5, 0.25, 0.01, help="规建类方案在总评分中的权重")
            wh_weight = st.slider("维护类权重", 0.0, 0.2, 0.1, 0.01, help="维护类方案在总评分中的权重")
            tt_weight = st.slider("天调类权重", 0.0, 0.3, 0.15, 0.01, help="天调类方案在总评分中的权重")
            zz_weight = st.slider("整治类权重", 0.0, 0.5, 0.25, 0.01, help="整治类方案在总评分中的权重")
        
        with col2:
            st.markdown("#### 评估设置")
            include_24_legacy = st.checkbox("包含24年遗留网格", value=True, help="是否包含24年遗留网格的评估")
            include_25_supervision = st.checkbox("包含25年督办网格", value=True, help="是否包含25年督办网格的评估")
            filter_fath = st.checkbox("过滤FATH方案", value=True, help="是否过滤包含_FATH_的方案ID")
        
        if st.button("开始网格评估", type="primary"):
            self._perform_grid_evaluation({
                'kc_score': kc_score,
                'gj_weight': gj_weight,
                'wh_weight': wh_weight,
                'tt_weight': tt_weight,
                'zz_weight': zz_weight,
                'include_24_legacy': include_24_legacy,
                'include_25_supervision': include_25_supervision,
                'filter_fath': filter_fath
            })
    
    def _render_grid_analysis(self):
        """渲染网格分析页面"""
        st.markdown("#### 📈 网格分析")
        st.info("选择网格和时间周期，查看结果分和过程分的雷达图分析")
        
        # 获取可用的时间周期
        available_periods = self.db_manager.get_available_time_periods()
        
        if not available_periods:
            st.warning("暂无网格结果得分数据，请先导入数据")
            return
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # 网格选择方式
            search_type = st.radio(
                "网格选择方式",
                ["按网格ID", "按网格中文名"],
                horizontal=True
            )
        
        with col2:
            if search_type == "按网格ID":
                # 获取所有网格ID
                all_grids = self._get_all_grids_from_results(available_periods)
                grid_ids = sorted(list(set([g['grid_code'] for g in all_grids])))
                selected_grid = st.selectbox("选择网格ID", grid_ids)
                grid_code = selected_grid
                grid_name = None
            else:
                # 获取所有网格中文名
                all_grids = self._get_all_grids_from_results(available_periods)
                grid_names = sorted(list(set([g['grid_name'] for g in all_grids if g.get('grid_name')])))
                selected_name = st.selectbox("选择网格中文名", grid_names)
                grid_code = None
                grid_name = selected_name
        
        with col3:
            # 粒度选择
            granularity = st.radio("粒度", ["周", "月"], horizontal=True)
            # 时间周期选择（根据粒度过滤）
            if granularity == "月":
                # 提取所有月份
                import re
                months = []
                for period in available_periods:
                    match = re.match(r'(\d{4}年\d+月)', period)
                    if match and match.group(1) not in months:
                        months.append(match.group(1))
                time_period = st.selectbox("选择月份", months, index=0 if months else None)
            else:
                time_period = st.selectbox("选择时间周期", available_periods, index=0)
        
        if st.button("开始分析", type="primary"):
            self._perform_grid_analysis(grid_code, grid_name, time_period, granularity)
    
    def _get_all_grids_from_results(self, time_periods):
        """从结果得分表获取所有网格信息"""
        all_grids_dict = {}  # 使用字典避免重复
        for period in time_periods[:10]:  # 只查询最近10个周期
            results = self.db_manager.get_grid_result_scores_by_time(period)
            for r in results:
                grid_code = r.get('grid_code')
                if grid_code and grid_code not in all_grids_dict:
                    all_grids_dict[grid_code] = {
                        'grid_code': grid_code,
                        'grid_name': r.get('grid_name'),
                        'city': r.get('city')
                    }
        return list(all_grids_dict.values())
    
    def _perform_grid_analysis(self, grid_code, grid_name, time_period, granularity):
        """执行网格分析"""
        try:
            # 获取网格结果得分数据
            if grid_code:
                # 如果是月粒度，需要获取该月的所有周数据并平均
                if granularity == "月":
                    # 提取年月（假设时间周期格式如"2025年6月第2周"）
                    month_data = self._get_month_data(grid_code, time_period)
                    if not month_data:
                        st.error(f"未找到该网格在{time_period}的数据")
                        return
                    result_data = month_data
                else:
                    result_data = self.db_manager.get_grid_result_score(grid_code, time_period)
            else:
                # 通过网格中文名查找
                if granularity == "月":
                    results = self.db_manager.get_grid_result_scores_by_time(time_period)
                    result_data = None
                    for r in results:
                        if r.get('grid_name') == grid_name:
                            month_data = self._get_month_data(r.get('grid_code'), time_period)
                            if month_data:
                                result_data = month_data
                                grid_code = r.get('grid_code')
                                break
                else:
                    results = self.db_manager.get_grid_result_scores_by_time(time_period)
                    result_data = None
                    for r in results:
                        if r.get('grid_name') == grid_name:
                            result_data = r
                            grid_code = r.get('grid_code')
                            break
            
            if not result_data:
                st.error("未找到该网格的数据")
                return
            
            # 解析结果数据JSON（月粒度已经平均，周粒度直接使用）
            if isinstance(result_data, dict) and 'grid_result_data_json' in result_data:
                grid_result_data_json = result_data.get('grid_result_data_json', '{}')
                if isinstance(grid_result_data_json, str):
                    result_json = json.loads(grid_result_data_json)
                else:
                    result_json = grid_result_data_json if grid_result_data_json else {}
            else:
                # 月粒度直接返回平均后的JSON
                result_json = result_data.get('averaged_data', {})
            
            # 计算结果分（5部分）
            result_scores = self._calculate_result_scores(result_data, result_json)
            
            # 计算过程分（6项）
            process_scores = self._calculate_process_scores(grid_code)
            
            # 获取质量得分（最终得分）
            quality_score = result_data.get('final_score')
            if quality_score is None:
                quality_score = 0.0
            else:
                try:
                    quality_score = float(quality_score)
                except (ValueError, TypeError):
                    quality_score = 0.0
            
            # 计算总过程得分
            total_process_score = self._calculate_total_process_score(grid_code)
            
            # 显示分析结果
            self._display_grid_analysis(result_data, result_scores, process_scores, quality_score, total_process_score, granularity)
            
        except Exception as e:
            st.error(f"网格分析失败: {str(e)}")
            self.logger.error(f"网格分析失败: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _get_month_data(self, grid_code, time_period):
        """获取月份的平均数据（将周粒度数据平均）"""
        # 从时间周期中提取年月（如"2025年6月第2周" -> "2025年6月"）
        import re
        match = re.match(r'(\d{4}年\d+月)', time_period)
        if not match:
            return None
        
        month_str = match.group(1)
        
        # 获取该月的所有周数据
        all_periods = self.db_manager.get_available_time_periods()
        month_periods = [p for p in all_periods if month_str in p]
        
        if not month_periods:
            return None
        
        # 获取该网格在该月的所有周数据
        week_data_list = []
        for period in month_periods:
            data = self.db_manager.get_grid_result_score(grid_code, period)
            if data:
                week_data_list.append(data)
        
        if not week_data_list:
            return None
        
        # 平均所有周的得分数据
        averaged_data = {}
        score_keys = []  # 收集所有得分字段
        
        for week_data in week_data_list:
            grid_result_data_json = week_data.get('grid_result_data_json', '{}')
            if isinstance(grid_result_data_json, str):
                week_json = json.loads(grid_result_data_json)
            else:
                week_json = grid_result_data_json if grid_result_data_json else {}
            
            for key, value in week_json.items():
                if '得分' in key or '扣分' in key:
                    if key not in score_keys:
                        score_keys.append(key)
                    if key not in averaged_data:
                        averaged_data[key] = []
                    try:
                        averaged_data[key].append(float(value) if value is not None else 0)
                    except (ValueError, TypeError):
                        pass
        
        # 计算平均值
        for key in score_keys:
            if averaged_data[key]:
                averaged_data[key] = sum(averaged_data[key]) / len(averaged_data[key])
            else:
                averaged_data[key] = 0
        
        # 使用第一周的基础信息，更新为月粒度
        base_data = week_data_list[0].copy()
        base_data['time_period'] = month_str
        base_data['averaged_data'] = averaged_data
        
        return base_data
    
    def _calculate_result_scores(self, result_data, result_json):
        """计算结果分（5部分）"""
        scores = {
            '投诉支撑': 0.0,
            '业务感知': 0.0,
            '网络基础': 0.0,
            '领先竞对': 0.0,
            '网络演进': 0.0
        }
        
        scene_area = result_data.get('scene_area_attribute', '')
        is_urban = scene_area in ['核心城区', '非核心城区']
        
        # 1. 投诉支撑
        complaint_score = self._get_score(result_json, '网格万投比得分')
        heavy_complaint_score = self._get_score(result_json, '网格中重投栅格数得分')
        scores['投诉支撑'] = (complaint_score + heavy_complaint_score) / 2
        
        # 2. 业务感知
        data_perception = (
            self._get_score(result_json, '4G下行RTT时延超300ms次数占比%得分') +
            self._get_score(result_json, '5G下行RTT时延超300ms次数占比%得分') +
            self._get_score(result_json, '4G下行RTT时延超300ms次数差小区占比%得分') +
            self._get_score(result_json, '5G下行RTT时延超300ms次数差小区占比%得分') +
            self._get_score(result_json, '4G视频质差话单占比%得分') +
            self._get_score(result_json, '5G视频质差话单占比%得分') +
            self._get_score(result_json, '4G视频质差小区占比%得分') +
            self._get_score(result_json, '5G视频质差小区占比%得分')
        ) / 8
        
        voice_perception = (
            self._get_score(result_json, '4G小区无线掉话率%得分') +
            self._get_score(result_json, '5G小区无线掉话率%得分') +
            self._get_score(result_json, '4G无线高掉话差小区占比%得分') +
            self._get_score(result_json, '5G无线高掉话差小区占比%得分') +
            self._get_score(result_json, '4G语音通话质差比例%得分') +
            self._get_score(result_json, 'VONR_语音通话质差比例%得分') +
            self._get_score(result_json, '4G语音通话质差小区占比%得分') +
            self._get_score(result_json, 'VONR_语音通话质差小区占比%得分')
        ) / 8
        
        scores['业务感知'] = (data_perception + voice_perception) / 2
        
        # 3. 网络基础
        # 覆盖
        if is_urban:
            coverage_score = (
                self._get_score(result_json, '4G问题楼宇总面积占比得分') +
                self._get_score(result_json, '5G问题楼宇总面积占比得分')
            ) / 2
        else:
            coverage_score = (
                self._get_score(result_json, '4G弱覆盖栅格占比（采样点>=100)%得分') +
                self._get_score(result_json, '5G弱覆盖栅格占比（采样点>=20)%得分')
            ) / 2
        
        # 容量
        capacity_score = (
            self._get_score(result_json, '5G高负荷待扩容问题小区占比%得分') +
            self._get_score(result_json, '4G高负荷待扩容问题小区占比%得分')
        ) / 2
        
        # 干扰
        interference_score = (
            self._get_score(result_json, '4G高干扰问题小区占比%得分') +
            self._get_score(result_json, '5G高干扰问题小区占比%得分')
        ) / 2
        
        # 结构
        structure_score = (
            self._get_score(result_json, '4G结构问题小区占比%得分') +
            self._get_score(result_json, '5G结构问题小区占比%得分')
        ) / 2
        
        scores['网络基础'] = (coverage_score + capacity_score + interference_score + structure_score) / 4
        
        # 4. 领先竞对
        if is_urban:
            competition_score = (
                self._get_score(result_json, '4G移动质差且劣于竞对楼宇面积占比%得分') +
                self._get_score(result_json, '5G移动质差且劣于竞对楼宇面积占比%得分')
            ) / 2
        else:
            competition_score = (
                self._get_score(result_json, '移动4G弱覆盖且劣于竞对栅格占比%得分') +
                self._get_score(result_json, '移动5G弱覆盖且劣于竞对栅格占比%得分')
            ) / 2
        
        scores['领先竞对'] = competition_score
        
        # 5. 网络演进
        flow_back_score = self._get_score(result_json, '5G倒流流量比%得分')
        # 4G城区宏站单D站点扣分（扣分值，需要转换为得分：扣分越少得分越高）
        penalty = self._get_score(result_json, '4G城区宏站单D站点扣分', 0)
        penalty_score = max(0, 100 - abs(penalty))  # 扣分转换为得分（扣分0分则得100分）
        
        scores['网络演进'] = (flow_back_score + penalty_score) / 2
        
        return scores
    
    def _get_score(self, result_json, key, default=0.0):
        """从结果JSON中获取得分值"""
        value = result_json.get(key, default)
        try:
            return float(value) if value is not None else default
        except (ValueError, TypeError):
            return default
    
    def _calculate_process_scores(self, grid_code):
        """计算过程分（6项）"""
        scores = {
            '规建类': 0.0,
            '天调类': 0.0,
            '维护类': 0.0,
            '整治类': 0.0,
            '时效性': 0.0,
            '准确性': 0.0
        }
        
        # 获取网格的面板数据
        grid_data = self.db_manager.get_panel_data_by_grid(grid_code)
        
        if not grid_data:
            return scores
        
        # 按方案类型统计（与_calculate_grid_process_score逻辑一致）
        scheme_types = {'规建': {'total': 0, 'success': 0}, 
                       '天调': {'total': 0, 'success': 0},
                       '维护': {'total': 0, 'success': 0},
                       '整治': {'total': 0, 'success': 0}}
        
        excluded_count = 0
        timeout_count = 0
        total_valid_schemes = 0  # 用于计算时效性的总方案数（排除剔除和方案剔除的）
        
        for record in grid_data:
            scheme_type = self._judge_scheme_type(record.get('scheme_type', ''), record.get('scheme', ''))
            vcisvail = record.get('vcisvail', '')
            scheme_id = record.get('scheme_id', '')
            
            # 跳过FATH方案
            if scheme_id and '_FATH_' in scheme_id:
                continue
            
            # 跳过24年完成的方案ID（不参与过程分计算）
            if scheme_id and self.db_manager.is_scheme_excluded(scheme_id):
                continue
            
            if scheme_type in ['规建', '天调', '维护', '整治']:
                # 只有非剔除状态的方案才计入总数
                # 注意：对于规建、天调、维护、整治这四类，"线下已完成"的方案应该计入准确性分母和完成数
                # 对于其他类型（优化、非无线等），"线下已完成"不计入准确性计算
                if vcisvail not in ['剔除', '方案剔除']:
                    scheme_types[scheme_type]['total'] += 1
                    total_valid_schemes += 1
                    
                    # 判断是否完成（统计"成功"和"线下已完成"）
                    # 因为这里已经是这四类方案，所以"线下已完成"应该计入完成数
                    implement_results = record.get('implement_results', '') or record.get('vcimplement_results', '')
                    if implement_results == '成功' or vcisvail == '线下已完成':
                        scheme_types[scheme_type]['success'] += 1
                
                # 统计剔除方案（用于准确性计算）
                # 注意：剔除方案包括：剔除、方案剔除、方案变更，以及线下已完成（仅限这四类）
                if vcisvail in ['剔除', '方案剔除', '方案变更', '线下已完成']:
                    excluded_count += 1
        
        # 计算各方案类型的完成率（转换为0-100分）
        for scheme_type in ['规建', '天调', '维护', '整治']:
            total = scheme_types[scheme_type]['total']
            success = scheme_types[scheme_type]['success']
            if total > 0:
                scores[f'{scheme_type}类'] = (success / total) * 100
            else:
                scores[f'{scheme_type}类'] = 0.0
        
        # 获取超时方案数量（与_calculate_grid_process_score逻辑一致）
        timeout_query = """
        SELECT COUNT(*) as count
        FROM timeout_scheme_list t
        INNER JOIN panel_data p ON t.scheme_id = p.scheme_id
        WHERE p.grid_code = ? 
        AND t.scheme_status IN ('已超时未完成', '超时已完成')
        AND (t.exclude_status IS NULL OR t.exclude_status NOT IN ('方案剔除', '剔除', '线下已完成', '方案变更'))
        """
        timeout_result = self.db_manager.execute_query(timeout_query, (grid_code,))
        timeout_count = timeout_result[0]['count'] if timeout_result else 0
        
        # 计算时效性得分（超时比例越低得分越高）
        # 总方案数需要统计所有非剔除的方案（包括优化类、非无线类等）
        # 排除FATH方案和24年完成的方案ID
        all_schemes_count = sum(1 for r in grid_data 
                                if r.get('scheme_id') 
                                and '_FATH_' not in r.get('scheme_id', '')
                                and not self.db_manager.is_scheme_excluded(r.get('scheme_id', ''))
                                and r.get('vcisvail') not in ['剔除', '方案剔除'])
        
        if all_schemes_count > 0:
            timeout_ratio = timeout_count / all_schemes_count
            if timeout_ratio < 0.20:
                scores['时效性'] = 100.0
            elif timeout_ratio > 0.6:
                scores['时效性'] = 0.0
            else:
                scores['时效性'] = (1 - (timeout_ratio - 0.20) / 0.40) * 100
        else:
            scores['时效性'] = 0.0
        
        # 计算准确性得分（剔除比例越低得分越高）
        accuracy_denominator = sum(s['total'] for s in scheme_types.values())
        if accuracy_denominator > 0:
            exclude_ratio = excluded_count / (accuracy_denominator + excluded_count)
            if exclude_ratio < 0.15:
                scores['准确性'] = 100.0
            elif exclude_ratio > 0.3:
                scores['准确性'] = 0.0
            else:
                scores['准确性'] = (1 - (exclude_ratio - 0.15) / 0.15) * 100
        else:
            scores['准确性'] = 0.0
        
        # 检查是否有勘察数据（如果没有任何方案类型数据，则全为0）
        has_survey = any(s['total'] > 0 for s in scheme_types.values())
        if not has_survey:
            return {k: 0.0 for k in scores}
        
        return scores
    
    def _calculate_total_process_score(self, grid_code):
        """计算网格的总过程得分"""
        # 获取网格的面板数据
        grid_data = self.db_manager.get_panel_data_by_grid(grid_code)
        
        if not grid_data:
            return 0.0
        
        # 获取网格标签（用于判断是否为2025年网格）
        grid_label = ''
        if grid_data:
            grid_label = grid_data[0].get('label', '')
        
        # 统计剔除和线下完成数量
        excluded_count = 0
        offline_completed_count = 0
        
        for record in grid_data:
            scheme_type = self._judge_scheme_type(record.get('scheme_type', ''), record.get('scheme', ''))
            vcisvail = record.get('vcisvail', '')
            scheme_id = record.get('scheme_id', '')
            
            if scheme_id and '_FATH_' in scheme_id:
                continue
            
            # 跳过24年完成的方案ID（不参与过程分计算）
            if scheme_id and self.db_manager.is_scheme_excluded(scheme_id):
                continue
            
            if scheme_type in ['规建', '维护', '天调', '整治']:
                # 统计剔除方案（用于准确性计算）：包括剔除、方案剔除、方案变更，以及线下已完成
                if vcisvail in ['剔除', '方案剔除', '方案变更', '线下已完成']:
                    excluded_count += 1
                if vcisvail == '线下已完成':
                    offline_completed_count += 1
        
        # 获取超时方案数量
        timeout_query = """
        SELECT COUNT(*) as count
        FROM timeout_scheme_list t
        INNER JOIN panel_data p ON t.scheme_id = p.scheme_id
        WHERE p.grid_code = ? 
        AND t.scheme_status IN ('已超时未完成', '超时已完成')
        AND (t.exclude_status IS NULL OR t.exclude_status NOT IN ('方案剔除', '剔除', '线下已完成', '方案变更'))
        """
        timeout_result = self.db_manager.execute_query(timeout_query, (grid_code,))
        timeout_count = timeout_result[0]['count'] if timeout_result else 0
        
        # 计算准确性分母
        accuracy_denominator = 0
        for record in grid_data:
            scheme_type = self._judge_scheme_type(record.get('scheme_type', ''), record.get('scheme', ''))
            vcisvail = record.get('vcisvail', '')
            # 排除剔除、方案剔除的方案，包含"线下已完成"的方案
            if vcisvail not in ['剔除', '方案剔除']:
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    accuracy_denominator += 1
        
        # 使用_calculate_grid_process_score计算总过程得分
        grid_score, score_details = self._calculate_grid_process_score(
            grid_data,
            grid_label,
            excluded_count,
            offline_completed_count,
            timeout_count,
            accuracy_denominator
        )
        
        # 转换为0-100分
        return grid_score * 100
    
    def _display_grid_analysis(self, result_data, result_scores, process_scores, quality_score, total_process_score, granularity="周"):
        """显示网格分析结果（雷达图）"""
        # 显示网格基本信息
        st.markdown("---")
        col1, col2, col3, col4, col5, col6 = st.columns(6)
        with col1:
            st.metric("网格ID", result_data.get('grid_code', ''))
        with col2:
            st.metric("网格名称", result_data.get('grid_name', ''))
        with col3:
            st.metric("地市", result_data.get('city', ''))
        with col4:
            time_label = f"{result_data.get('time_period', '')} ({granularity}粒度)"
            st.metric("时间周期", time_label)
        with col5:
            st.metric("质量得分", f"{quality_score:.2f}")
        with col6:
            st.metric("过程得分", f"{total_process_score:.2f}")
        
        # 创建左右两列的雷达图
        col_left, col_right = st.columns(2)
        
        with col_left:
            st.markdown("#### 📊 结果分析")
            fig_result = self._create_radar_chart(
                result_scores,
                categories=list(result_scores.keys()),
                title="结果分分析",
                max_value=100
            )
            st.pyplot(fig_result)
            plt.close(fig_result)
            
            # 显示详细得分
            with st.expander("查看详细得分"):
                for category, score in result_scores.items():
                    st.write(f"**{category}**: {score:.2f}分")
        
        with col_right:
            st.markdown("#### 📈 过程分析")
            # 检查是否有勘察数据
            has_survey = any(v > 0 for v in process_scores.values())
            if has_survey:
                fig_process = self._create_radar_chart(
                    process_scores,
                    categories=list(process_scores.keys()),
                    title="过程分分析",
                    max_value=100
                )
                st.pyplot(fig_process)
                plt.close(fig_process)
            else:
                # 没有勘察数据，显示提示
                st.info("该网格暂无方案数据（未勘察）")
                # 仍然绘制全为0的雷达图
                fig_process = self._create_radar_chart(
                    process_scores,
                    categories=list(process_scores.keys()),
                    title="过程分分析（未勘察）",
                    max_value=100
                )
                st.pyplot(fig_process)
                plt.close(fig_process)
            
            # 显示详细得分
            with st.expander("查看详细得分"):
                for category, score in process_scores.items():
                    st.write(f"**{category}**: {score:.2f}分")
    
    def _create_radar_chart(self, scores, categories, title, max_value=100):
        """创建雷达图"""
        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 角度
        num_vars = len(categories)
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
        angles += angles[:1]  # 闭合
        
        # 数据
        values = [scores.get(cat, 0) for cat in categories]
        values += values[:1]  # 闭合
        
        # 创建图形
        fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(projection='polar'))
        
        # 绘制
        ax.plot(angles, values, 'o-', linewidth=2, label=title)
        ax.fill(angles, values, alpha=0.25)
        
        # 设置标签
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories, fontsize=10)
        
        # 设置y轴范围
        ax.set_ylim(0, max_value)
        ax.set_yticks([20, 40, 60, 80, 100])
        ax.set_yticklabels(['20', '40', '60', '80', '100'], fontsize=8)
        ax.grid(True)
        
        plt.title(title, size=14, fontweight='bold', pad=20)
        
        return fig
    
    def _render_scheme_analysis(self):
        """渲染方案分析页面"""
        st.subheader("📈 方案分析")
        
        # 分析类型选择
        analysis_type = st.radio(
            "选择分析类型",
            ["地市汇总分析", "网格明细分析", "方案类型分析", "实施结果分析"],
            horizontal=True
        )
        
        if analysis_type == "地市汇总分析":
            self._render_city_summary_analysis()
        elif analysis_type == "网格明细分析":
            self._render_grid_detail_analysis()
        elif analysis_type == "方案类型分析":
            self._render_scheme_type_analysis()
        elif analysis_type == "实施结果分析":
            self._render_implementation_analysis()
    
    def _render_data_export(self):
        """渲染数据导出页面"""
        st.subheader("📤 数据导出")
        
        # 导出选项
        export_options = st.multiselect(
            "选择导出内容",
            ["地市汇总评估过程得分", "网格明细评估过程得分", "方案实施统计", "网格标签统计", "未完成的方案"],
            default=["地市汇总评估过程得分", "网格明细评估过程得分"]
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            if export_options:
                if st.button("生成导出文件", type="primary"):
                    self._export_data(export_options)
        
        with col2:
            # 检查是否有缓存的导出文件
            if 'export_file_data' in st.session_state and 'export_filename' in st.session_state:
                try:
                    # 检查文件是否存在
                    if os.path.exists(st.session_state['export_file_data']):
                        with open(st.session_state['export_file_data'], "rb") as file:
                            file_data = file.read()
                            st.download_button(
                                label="📥 下载导出文件",
                                data=file_data,
                                file_name=st.session_state['export_filename'],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary"
                            )
                    else:
                        st.warning("导出文件已过期，请重新生成")
                        # 清理session_state
                        if 'export_file_data' in st.session_state:
                            del st.session_state['export_file_data']
                        if 'export_filename' in st.session_state:
                            del st.session_state['export_filename']
                except Exception as e:
                    st.error(f"读取导出文件失败: {str(e)}")
                    # 清理session_state
                    if 'export_file_data' in st.session_state:
                        del st.session_state['export_file_data']
                    if 'export_filename' in st.session_state:
                        del st.session_state['export_filename']
            else:
                st.info("请先生成导出文件")
    
    def _process_uploaded_files(self, uploaded_files):
        """处理上传的文件"""
        try:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            total_files = len(uploaded_files)
            processed_data = []
            
            # 生成批次ID
            batch_id = f"panel_import_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4())[:8]}"
            
            # 清空旧数据（每次导入前清空）
            if st.checkbox("清空旧数据", value=True, help="每次导入新数据前清空数据库中的旧数据"):
                if self.db_manager.clear_panel_data():
                    st.info("✅ 旧数据已清空")
                else:
                    st.error("❌ 清空旧数据失败")
                    return
            
            # 创建导入批次记录
            self.db_manager.create_panel_import_batch(batch_id, total_files, f"面板数据导入_{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            for i, file in enumerate(uploaded_files):
                status_text.text(f"正在处理文件 {i + 1}/{total_files}: {file.name}")
                
                try:
                    # 读取文件（支持CSV和Excel格式）
                    df = self._read_excel_or_csv(file)
                    
                    # 处理数据
                    file_data = self._process_csv_data(df, file.name)
                    processed_data.extend(file_data)
                    
                    st.success(f"✅ 文件 {file.name} 处理完成，共 {len(file_data)} 条记录")
                    
                except Exception as e:
                    st.error(f"❌ 文件 {file.name} 处理失败: {str(e)}")
                    self.logger.error(f"文件处理失败: {str(e)}")
                    import traceback
                    self.logger.error(traceback.format_exc())
                    continue
                
                # 更新进度
                progress = (i + 1) / total_files
                progress_bar.progress(progress)
            
            # 保存处理后的数据到数据库
            if processed_data:
                # 批量插入数据库
                if self.db_manager.insert_panel_data(processed_data, batch_id):
                    # 更新批次统计
                    self.db_manager.update_panel_import_batch(
                        batch_id, 
                        len(processed_data), 
                        len(processed_data), 
                        0, 
                        "completed"
                    )
                    
                    # 保存到session state
                    st.session_state['processed_panel_data'] = processed_data
                    st.session_state['current_batch_id'] = batch_id
                    
                    st.success(f"🎉 数据处理完成！共处理 {len(processed_data)} 条记录")
                    st.success(f"📊 数据已保存到数据库，批次ID: {batch_id}")
                    
                    # 显示数据预览
                    with st.expander("查看数据预览"):
                        preview_df = pd.DataFrame(processed_data[:10])  # 显示前10条
                        st.dataframe(preview_df)
                    
                    # 显示数据库统计
                    summary = self.db_manager.get_panel_data_summary(batch_id)
                    if summary:
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("总记录数", summary.get('total_records', 0))
                        with col2:
                            st.metric("地市数", summary.get('city_count', 0))
                        with col3:
                            st.metric("网格数", summary.get('grid_count', 0))
                        with col4:
                            st.metric("方案类型数", summary.get('scheme_type_count', 0))
                else:
                    st.error("❌ 数据保存到数据库失败")
            else:
                st.warning("⚠️ 没有有效的数据被处理")
                
        except Exception as e:
            st.error(f"文件处理过程失败: {str(e)}")
            self.logger.error(f"文件处理失败: {str(e)}")
    
    def _process_csv_data(self, df, filename):
        """处理CSV数据，支持两种格式：total文件和超时方案清单文件"""
        processed_data = []
        
        try:
            # 判断文件格式：检查是否包含超时方案清单的标准字段
            is_timeout_format = '方案提交时间' in df.columns or '方案完成时间' in df.columns
            
            # 列名映射：total文件列名 -> 超时文件列名（英文字段名）
            column_mapping = {
                # total文件列名 -> 超时文件列名（英文字段名）
                'vcorder_code': 'order_number',  # 工单编号
                'business_key_': 'optimize_number',  # 优化单号
                'current_act_name': 'process_status',  # 流程状态
                'start_time_': 'start_time',  # 开始时间
                'vcsmall_grid_code': 'grid_code',  # 微网格编号
                'vcsmall_grid_name': 'grid_name',  # 微网格名称
                'vclabel': 'label',  # 标签
                'vccity': 'city',  # 地市
                'vcdistrict': 'district',  # 区县
                'vcreasontype': 'reason_category',  # 原因分类
                'vcreason': 'root_cause',  # 根本原因
                'vcschemetype': 'scheme_category',  # 方案分类
                'vcscheme': 'measures',  # 措施
                'vcsonorder_type': 'scheme_type',  # 方案类型
                'vcoptimize_object_name': 'cell_name',  # 小区名称
                'vcadjust_parameters': 'adjust_param',  # 调整参数
                'vcadjust_before_value': 'adjust_before_value',  # 调整前值
                'vcadjust_target_value': 'target_value',  # 目标值
                'vcmeasure_code': 'sub_order_number',  # 子工单号
                'vcorder_status': 'sub_order_status',  # 子工单状态
                'vcimplement_results': 'implement_results',  # 实施结果
                'vcscheme_id': 'scheme_id',  # 方案id
                'vcisvail': 'exclude_status',  # 剔除/线下完成
                'update_label': 'update_label',  # 更新标签
            }
            
            # 超时方案清单格式的中文列名映射到英文字段名
            timeout_column_mapping = {
                '工单编号': 'order_number',
                '优化单号': 'optimize_number',
                '流程状态': 'process_status',
                '开始时间': 'start_time',
                '微网格编号': 'grid_code',
                '微网格名称': 'grid_name',
                '标签': 'label',
                '地市': 'city',
                '区县': 'district',
                '原因分类': 'reason_category',
                '根本原因': 'root_cause',
                '方案分类': 'scheme_category',
                '措施': 'measures',
                '方案类型': 'scheme_type',
                '小区名称': 'cell_name',
                '调整参数': 'adjust_param',
                '调整前值': 'adjust_before_value',
                '目标值': 'target_value',
                '子工单号': 'sub_order_number',
                '子工单状态': 'sub_order_status',
                '实施结果': 'implement_results',
                '方案id': 'scheme_id',
                '剔除/线下完成': 'exclude_status',
                '更新标签': 'update_label',
                '方案提交时间': 'scheme_submit_time',
                '方案完成时间': 'scheme_complete_time',
                '方案执行耗时': 'scheme_execution_time',
                '方案标准时长': 'scheme_standard_time',
                '方案状态': 'scheme_status',
            }
            
            for num in range(len(df)):
                record = {}
                
                if is_timeout_format:
                    # 超时方案清单格式：直接使用中文列名映射到英文字段名
                    for chinese_col, english_field in timeout_column_mapping.items():
                        if chinese_col in df.columns:
                            value = df[chinese_col].iloc[num]
                            # 处理数值字段
                            if english_field in ['scheme_execution_time', 'scheme_standard_time']:
                                try:
                                    if pd.notna(value) and value != '' and str(value).strip() != 'nan':
                                        record[english_field] = float(value)
                                    else:
                                        record[english_field] = None
                                except (ValueError, TypeError):
                                    record[english_field] = None
                            else:
                                record[english_field] = str(value).strip() if pd.notna(value) else ''
                        else:
                            record[english_field] = ''
                    
                    # 获取标签字段进行过滤
                    label = record.get('label', '')
                    
                    # 判断是否为24年网格：
                    # 必须同时包含"2025年督办微网格"和24年标签（2024年第二批督办微网格、2024年遗留、2024年第一批督办重要场景所在网格）
                    # 如果单独只有24年标签，不导入
                    is_2024_grid = ('2025年督办微网格' in label and 
                                   ('2024年第二批督办微网格' in label or 
                                    '2024年遗留' in label or 
                                    '2024年第一批督办重要场景所在网格' in label))
                    
                    # 判断是否为25年网格：
                    # 标签单独只有"2025年督办微网格"，不包含任何24年标签
                    is_2025_grid = ('2025年督办微网格' in label and 
                                   '2024年第二批督办微网格' not in label and
                                   '2024年遗留' not in label and
                                   '2024年第一批督办重要场景所在网格' not in label)
                    
                    # 过滤条件：只保留24年网格或25年督办微网格的数据
                    # 24年网格：必须同时包含"2025年督办微网格"和24年标签（单独只有24年标签不导入）
                    # 25年网格：只包含"2025年督办微网格"，不包含任何24年标签
                    if not (is_2024_grid or is_2025_grid):
                        continue  # 跳过不符合条件的记录
                    
                    # 保留兼容字段
                    record['vcoptimize_object_name'] = record.get('cell_name', '')
                    record['vcisvail'] = record.get('exclude_status', '')
                    record['vcmeasure_code'] = record.get('sub_order_number', '')
                    record['current_act_name'] = record.get('process_status', '')
                    
                else:
                    # total文件格式：需要映射列名并检查过滤条件
                    # 检查必要字段
                    if 'current_act_name' not in df.columns or 'vcisvail' not in df.columns:
                        continue
                    
                    # 过滤条件 - 保留所有数据，包括剔除和方案剔除的数据
                    current_act_name = str(df['current_act_name'].iloc[num]).strip() if pd.notna(df['current_act_name'].iloc[num]) else ''
                    update_label = str(df.get('update_label', '').iloc[num]).strip() if 'update_label' in df.columns and pd.notna(df['update_label'].iloc[num]) else ''
                    vclabel = str(df.get('vclabel', '').iloc[num]).strip() if 'vclabel' in df.columns and pd.notna(df['vclabel'].iloc[num]) else ''
                    
                    # 判断是否为24年网格：
                    # 必须同时包含"2025年督办微网格"和24年标签（2024年第二批督办微网格、2024年遗留、2024年第一批督办重要场景所在网格）
                    # 如果单独只有24年标签，不导入
                    is_2024_grid = ('2025年督办微网格' in vclabel and 
                                   ('2024年第二批督办微网格' in vclabel or 
                                    '2024年遗留' in vclabel or 
                                    '2024年第一批督办重要场景所在网格' in vclabel))
                    
                    # 判断是否为25年网格：
                    # 标签单独只有"2025年督办微网格"，不包含任何24年标签
                    is_2025_grid = ('2025年督办微网格' in vclabel and 
                                   '2024年第二批督办微网格' not in vclabel and
                                   '2024年遗留' not in vclabel and
                                   '2024年第一批督办重要场景所在网格' not in vclabel)
                    
                    # 过滤条件：只保留24年网格或25年督办微网格的数据
                    # 24年网格：必须同时包含"2025年督办微网格"和24年标签（单独只有24年标签不导入）
                    # 25年网格：只包含"2025年督办微网格"，不包含任何24年标签
                    if (current_act_name == '方案实施结果' and 
                        update_label != '24年年底已实施完成' and
                        (is_2024_grid or is_2025_grid)):
                        
                        # 映射total文件的列名到英文字段名
                        for total_col, english_field in column_mapping.items():
                            if total_col in df.columns:
                                value = df[total_col].iloc[num]
                                record[english_field] = str(value).strip() if pd.notna(value) else ''
                            else:
                                record[english_field] = ''
                        
                        # 超时方案清单特有的5个字段设为空
                        record['scheme_submit_time'] = ''
                        record['scheme_complete_time'] = ''
                        record['scheme_execution_time'] = None
                        record['scheme_standard_time'] = None
                        record['scheme_status'] = ''
                        
                        # 保留兼容字段
                        record['vcoptimize_object_name'] = record.get('cell_name', '')
                        record['vcisvail'] = record.get('exclude_status', '')
                        record['vcmeasure_code'] = record.get('sub_order_number', '')
                        record['current_act_name'] = current_act_name
                
                # 添加文件名
                record['filename'] = filename
                
                # 检查必要字段是否存在
                if record.get('grid_code') and record.get('scheme_id'):
                    # 过滤掉24年完成的方案ID（不参与过程分计算）
                    scheme_id = record.get('scheme_id', '')
                    if scheme_id and self.db_manager.is_scheme_excluded(scheme_id):
                        continue  # 跳过这个方案，不导入
                    processed_data.append(record)
        
        except Exception as e:
            self.logger.error(f"处理CSV数据失败: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
        
        return processed_data
    
    def _perform_grid_evaluation(self, params):
        """执行网格评估"""
        try:
            # 检查数据库中是否有面板数据
            summary = self.db_manager.get_panel_data_summary()
            if not summary or summary.get('total_records', 0) == 0:
                st.error("❌ 数据库中没有面板数据，请先导入数据")
                return
            
            st.info(f"📊 数据库中共有 {summary.get('total_records', 0)} 条记录，开始评估...")
            
            # 从数据库获取数据
            data = self.db_manager.execute_query("SELECT * FROM panel_data ORDER BY city, grid_code")
            
            # 初始化评估结果
            evaluation_results = self._evaluate_grids(data, params)
            
            # 显示评估结果
            self._display_evaluation_results(evaluation_results)
            
        except Exception as e:
            st.error(f"网格评估失败: {str(e)}")
            self.logger.error(f"网格评估失败: {str(e)}")
    
    def _evaluate_grids(self, data, params):
        """评估网格数据（使用新公式）"""
        evaluation_results = {
            'city_summary': {},
            'grid_details': [],
            'scheme_statistics': {}
        }
        
        # 获取超时方案数据
        timeout_schemes_query = """
        SELECT DISTINCT p.scheme_id, p.grid_code, p.city, t.scheme_status, t.exclude_status
        FROM panel_data p
        INNER JOIN timeout_scheme_list t ON p.scheme_id = t.scheme_id
        WHERE t.scheme_status IN ('已超时未完成', '超时已完成')
        AND (t.exclude_status IS NULL OR t.exclude_status NOT IN ('方案剔除', '剔除', '线下已完成', '方案变更'))
        """
        timeout_schemes_data = self.db_manager.execute_query(timeout_schemes_query)
        
        # 按网格分组超时方案数据
        timeout_by_grid = {}
        for timeout_record in timeout_schemes_data:
            grid_key = f"{timeout_record['city']}_{timeout_record['grid_code']}"
            if grid_key not in timeout_by_grid:
                timeout_by_grid[grid_key] = 0
            timeout_by_grid[grid_key] += 1
        
        # 按城市和网格分组数据
        city_grid_data = {}
        for record in data:
            city = record['city']
            grid_code = record['grid_code']
            grid_key = f"{city}_{grid_code}"
            
            if city not in city_grid_data:
                city_grid_data[city] = {}
            if grid_code not in city_grid_data[city]:
                city_grid_data[city][grid_code] = {
                    'label': record.get('label', ''),
                    'schemes': [],
                    'excluded_count': 0,
                    'offline_completed_count': 0,
                    'timeout_count': 0
                }
            
            city_grid_data[city][grid_code]['schemes'].append(record)
            
            # 统计剔除和线下完成数量
            # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
            scheme_category = record.get('scheme_category', '') or record.get('scheme_type', '')
            measures = record.get('measures', '') or record.get('scheme', '')
            scheme_type = self._judge_scheme_type(scheme_category, measures)
            vcisvail = record.get('vcisvail', '')
            scheme_id = record.get('scheme_id', '')
            
            if scheme_id and '_FATH_' in scheme_id:
                continue
            
            # 统计剔除方案总数（用于准确性计算）：包括剔除、方案剔除、方案变更，以及线下已完成（仅限这四类）
            if vcisvail in ['剔除', '方案剔除', '方案变更', '线下已完成']:
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    city_grid_data[city][grid_code]['excluded_count'] += 1
            
            # 统计线下完成总数
            if vcisvail == '线下已完成':
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    city_grid_data[city][grid_code]['offline_completed_count'] += 1
        
        # 设置超时方案计数
        for city, grids in city_grid_data.items():
            for grid_code, grid_info in grids.items():
                grid_key = f"{city}_{grid_code}"
                grid_info['timeout_count'] = timeout_by_grid.get(grid_key, 0)
        
        # 计算每个网格的得分
        for city, grids in city_grid_data.items():
            city_stats = {
                'total_grids': len(grids),
                'total_score': 0,
                'scheme_types': {},
                'grids_2024': [],
                'grids_2025': []
            }
            
            for grid_code, grid_info in grids.items():
                # 计算准确性分母
                schemes_count = {}
                for scheme in grid_info['schemes']:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
                    measures = scheme.get('measures', '') or scheme.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    vcisvail = scheme.get('vcisvail', '')
                    # 排除剔除、方案剔除的方案
                    if vcisvail not in ['剔除', '方案剔除']:
                        if scheme_type in ['规建', '维护', '天调', '整治']:
                            if scheme_type not in schemes_count:
                                schemes_count[scheme_type] = 0
                            schemes_count[scheme_type] += 1
                
                accuracy_denominator = (
                    schemes_count.get('规建', 0) +
                    schemes_count.get('维护', 0) +
                    schemes_count.get('天调', 0) +
                    schemes_count.get('整治', 0)
                )
                
                # 使用新公式计算过程得分
                grid_score, score_details = self._calculate_grid_process_score(
                    grid_info['schemes'],
                    grid_info['label'],
                    grid_info['excluded_count'],
                    grid_info['offline_completed_count'],
                    grid_info['timeout_count'],
                    accuracy_denominator
                )
                
                scheme_stats = score_details.get('scheme_types', {})
                
                evaluation_results['grid_details'].append({
                    'city': city,
                    'grid_code': grid_code,
                    'score': grid_score,
                    'scheme_count': len(grid_info['schemes']),
                    'scheme_stats': scheme_stats
                })
                
                # 判断网格类型
                is_2024_grid = ('2024年第二批督办微网格' in grid_info['label'] or 
                              '2024年遗留' in grid_info['label'] or 
                              '2024年第一批督办重要场景所在网格' in grid_info['label'])
                is_2025_grid = '2025年督办微网格' in grid_info['label']
                
                if is_2024_grid:
                    city_stats['grids_2024'].append(grid_score)
                if is_2025_grid:
                    city_stats['grids_2025'].append(grid_score)
                
                city_stats['total_score'] += grid_score
                
                # 统计方案类型
                for scheme_type, stats in scheme_stats.items():
                    if scheme_type not in city_stats['scheme_types']:
                        city_stats['scheme_types'][scheme_type] = {'total': 0, 'success': 0}
                    city_stats['scheme_types'][scheme_type]['total'] += stats['total']
                    city_stats['scheme_types'][scheme_type]['success'] += stats['success']
            
            # 计算地市过程分：2024年平均分 + 2025年平均分
            avg_score_2024 = sum(city_stats['grids_2024']) / len(city_stats['grids_2024']) if city_stats['grids_2024'] else 0
            avg_score_2025 = sum(city_stats['grids_2025']) / len(city_stats['grids_2025']) if city_stats['grids_2025'] else 0
            city_stats['avg_score'] = avg_score_2024 + avg_score_2025
            
            evaluation_results['city_summary'][city] = city_stats
        
        return evaluation_results
    
    def _calculate_grid_score(self, schemes, params):
        """计算网格得分"""
        # 基于run_no_fath.py中的得分计算逻辑
        
        # 按方案类型分组
        scheme_types = {
            '维护': {'total': 0, 'success': 0, 'rate': 0.0},
            '优化': {'total': 0, 'success': 0, 'rate': 0.0},
            '规建': {'total': 0, 'success': 0, 'rate': 0.0},
            '整治': {'total': 0, 'success': 0, 'rate': 0.0},
            '天调': {'total': 0, 'success': 0, 'rate': 0.0},
            '非无线': {'total': 0, 'success': 0, 'rate': 0.0}
        }
        
        # 统计各类型方案
        for scheme in schemes:
            # 排除剔除、方案剔除的数据，不参与得分计算
            # 注意："线下已完成"应该参与得分计算
            if scheme.get('vcisvail') in ['剔除', '方案剔除']:
                continue
            
            # 排除包含_FATH_的方案（替换流程的工单）
            scheme_id = scheme.get('scheme_id', '')
            if scheme_id and '_FATH_' in scheme_id:
                continue
                
            # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
            scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
            measures = scheme.get('measures', '') or scheme.get('scheme', '')
            scheme_type = self._judge_scheme_type(scheme_category, measures)
            
            if scheme_type in scheme_types:
                scheme_types[scheme_type]['total'] += 1
                
                # 判断是否成功（统计"成功"和"线下已完成"）
                vcisvail = scheme.get('vcisvail', '')
                if scheme['implement_results'] == '成功' or vcisvail == '线下已完成':
                    scheme_types[scheme_type]['success'] += 1
        
        # 计算各类型完成率
        for scheme_type, stats in scheme_types.items():
            if stats['total'] > 0:
                stats['rate'] = stats['success'] / stats['total']
        
        return scheme_types
    
    def _calculate_grid_process_score(self, schemes, grid_label, excluded_count, offline_completed_count, timeout_count, accuracy_denominator):
        """计算网格过程分（新公式）"""
        # 判断网格类型：2024年或2025年
        # 优先判断2024年网格（如果包含2024年标签，即使同时包含2025年标签，也按2024年计算）
        is_2024_grid = ('2024年第二批督办微网格' in grid_label or 
                       '2024年遗留' in grid_label or 
                       '2024年第一批督办重要场景所在网格' in grid_label)
        # 只有在不是2024年网格的情况下，才判断为2025年网格
        is_2025_grid = not is_2024_grid and '2025年督办微网格' in grid_label
        
        # 统计各类型方案（用于计算方案过程分）
        scheme_types = {
            '维护': {'total': 0, 'success': 0},
            '规建': {'total': 0, 'success': 0},
            '整治': {'total': 0, 'success': 0},
            '天调': {'total': 0, 'success': 0}
        }
        
        # 统计子工单数量（用于计算勘察得分）
        sub_order_count = 0
        
        for scheme in schemes:
            # 排除包含_FATH_的方案（替换流程的工单）
            scheme_id = scheme.get('scheme_id', '')
            if scheme_id and '_FATH_' in scheme_id:
                continue
            
            # 统计子工单（current_act_name = "方案实施结果"）
            if scheme.get('current_act_name') == '方案实施结果':
                sub_order_count += 1
            
            # 判断方案类型
            # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
            scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
            measures = scheme.get('measures', '') or scheme.get('scheme', '')
            scheme_type = self._judge_scheme_type(scheme_category, measures)
            vcisvail = scheme.get('vcisvail', '')
            
            # 排除剔除、方案剔除的数据，不参与方案过程分计算
            # 注意："线下已完成"的方案：只有规建、天调、维护、整治这四类才计入准确性分母和完成数
            # scheme_types只包含这四类，所以这里的逻辑是正确的
            if vcisvail not in ['剔除', '方案剔除']:
                if scheme_type in scheme_types:
                    scheme_types[scheme_type]['total'] += 1
                    # 判断是否成功（统计"成功"和"线下已完成"）
                    # 因为scheme_types只包含规建、天调、维护、整治这四类，所以这里可以统计"线下已完成"
                    if scheme['implement_results'] == '成功' or vcisvail == '线下已完成':
                        scheme_types[scheme_type]['success'] += 1
        
        # 1. 计算勘察得分（5分）
        # 如果该网格有至少1个子工单（current_act_name = "方案实施结果"），得5分，否则0分
        kc_score = 5 if sub_order_count >= 1 else 0
        kc_weight = 5
        
        # 如果勘察得分为0，则后续都为0分
        if kc_score == 0:
            return 0.0, {
                'kc_score': 0,
                'scheme_process_score': 0,
                'accuracy_score': 0,
                'timeliness_score': 0,
                'scheme_types': scheme_types
            }
        
        # 2. 计算方案过程分（75分）
        scheme_process_score = 0.0
        scheme_process_weight = 0.0
        
        # 规建类 (25%)
        if scheme_types['规建']['total'] > 0:
            scheme_process_score += scheme_types['规建']['success'] / scheme_types['规建']['total'] * 25
            scheme_process_weight += 25
        
        # 维护类 (10%)
        if scheme_types['维护']['total'] > 0:
            scheme_process_score += scheme_types['维护']['success'] / scheme_types['维护']['total'] * 10
            scheme_process_weight += 10
        
        # 天调类 (15%)
        if scheme_types['天调']['total'] > 0:
            scheme_process_score += scheme_types['天调']['success'] / scheme_types['天调']['total'] * 15
            scheme_process_weight += 15
        
        # 整治类（结构整治类）(25%)
        if scheme_types['整治']['total'] > 0:
            scheme_process_score += scheme_types['整治']['success'] / scheme_types['整治']['total'] * 25
            scheme_process_weight += 25
        
        # 3. 计算方案准确性（10分）
        accuracy_score = 0.0
        accuracy_weight = 0.0
        
        if is_2025_grid:  # 2025年网格才计算准确性
            # 分母 = 准确性分母 + 剔除方案总数
            accuracy_total = accuracy_denominator + excluded_count
            if accuracy_total > 0:
                exclude_ratio = excluded_count / accuracy_total
                if exclude_ratio < 0.15:
                    accuracy_score = 10
                elif exclude_ratio > 0.3:
                    accuracy_score = 0
                else:
                    # 线性得分：(1-(剔除方案总数/(准确性分母+剔除方案总数)-0.15)/0.15)*10
                    accuracy_score = (1 - (exclude_ratio - 0.15) / 0.15) * 10
                accuracy_weight = 10
        
        # 4. 计算方案时效性（10分）
        timeliness_score = 0.0
        timeliness_weight = 0.0
        
        if is_2025_grid:  # 2025年网格才计算时效性
            # 总方案数 = 优化类总方案数 + 非无线类总方案数 + 规建类总方案数 + 维护类总方案数 + 天调类总方案数 + 整治类总方案数
            # 需要统计所有方案类型（排除剔除和方案剔除的）
            total_schemes = 0
            for scheme in schemes:
                scheme_id = scheme.get('scheme_id', '')
                vcisvail = scheme.get('vcisvail', '')
                # 排除包含_FATH_的方案
                if scheme_id and '_FATH_' in scheme_id:
                    continue
                # 排除剔除、方案剔除的数据
                # 注意："线下已完成"的方案应该计入总方案数
                if vcisvail not in ['剔除', '方案剔除']:
                    total_schemes += 1
            
            if total_schemes > 0:
                timeout_ratio = timeout_count / total_schemes
                if timeout_ratio < 0.20:
                    timeliness_score = 10
                elif timeout_ratio > 0.6:
                    timeliness_score = 0
                else:
                    # 线性得分：(1-(超时方案/总方案数-0.20)/0.40)*10
                    # 范围是0.20到0.6，所以是(0.6-0.20)=0.40的范围
                    timeliness_score = (1 - (timeout_ratio - 0.20) / 0.40) * 10
                timeliness_weight = 10
        
        # 5. 计算总过程分
        if is_2025_grid:
            # 2025年督办微网格计算方法
            total_score = kc_score + scheme_process_score + accuracy_score + timeliness_score
            total_weight = kc_weight + scheme_process_weight + accuracy_weight + timeliness_weight
        else:
            # 2024年网格计算方法
            total_score = kc_score + scheme_process_score
            total_weight = kc_weight + scheme_process_weight
        
        if total_weight > 0:
            process_score = total_score / total_weight
        else:
            process_score = 0.0
        
        return process_score, {
            'kc_score': kc_score,
            'scheme_process_score': scheme_process_score,
            'scheme_process_weight': scheme_process_weight,  # 方案过程分实际权重
            'accuracy_score': accuracy_score,
            'timeliness_score': timeliness_score,
            'total_weight': total_weight,
            'is_2025_grid': is_2025_grid,
            'scheme_types': scheme_types
        }
    
    def _judge_scheme_type(self, vcscheme_type, vcscheme):
        """判断方案类型"""
        if vcscheme_type == '非优化':
            return '非优化'
        elif vcscheme_type == '维护' and vcscheme == '天线方向角下倾角调整':
            return '天调'
        elif vcscheme_type == '维护' and vcscheme != '天线方向角下倾角调整':
            return '维护'
        elif vcscheme_type == '建设':
            return '规建'
        elif vcscheme_type == '规建':
            if vcscheme in ['新增宏微站建设', '新增室分建设']:
                return '规建'
            else:
                return '整治'
        else:
            return '优化'
    
    def _display_evaluation_results(self, results):
        """显示评估结果"""
        st.subheader("📊 评估结果")
        
        # 显示地市汇总
        if results['city_summary']:
            st.markdown("#### 地市汇总评估结果")
            city_data = []
            for city, stats in results['city_summary'].items():
                city_data.append({
                    '地市': city,
                    '网格数': stats['total_grids'],
                    '平均得分': round(stats['avg_score'] * 100, 2),
                    '总得分': round(stats['total_score'] * 100, 2)
                })
            
            if city_data:
                city_df = pd.DataFrame(city_data)
                st.dataframe(city_df)
                
                # 显示地市得分图表
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("##### 地市平均得分")
                    st.bar_chart(city_df.set_index('地市')['平均得分'])
                with col2:
                    st.markdown("##### 地市网格数")
                    st.bar_chart(city_df.set_index('地市')['网格数'])
        
        # 显示网格明细
        if results['grid_details']:
            st.markdown("#### 网格明细评估结果")
            
            # 创建网格明细表格
            grid_data = []
            for grid in results['grid_details']:
                grid_data.append({
                    '地市': grid['city'],
                    '网格ID': grid['grid_code'],
                    '过程得分': round(grid['score'] * 100, 2),
                    '方案数': grid['scheme_count']
                })
            
            if grid_data:
                grid_df = pd.DataFrame(grid_data)
                st.dataframe(grid_df)
                
                # 显示网格得分分布
                st.markdown("##### 网格得分分布")
                col1, col2 = st.columns(2)
                with col1:
                    st.histogram(grid_df['过程得分'], bins=20)
                with col2:
                    st.scatter_chart(grid_df[['方案数', '过程得分']])
        
        # 显示方案统计
        if results['city_summary']:
            st.markdown("#### 方案类型统计")
            
            # 汇总所有城市的方案统计
            total_scheme_stats = {}
            for city, stats in results['city_summary'].items():
                for scheme_type, type_stats in stats['scheme_types'].items():
                    if scheme_type not in total_scheme_stats:
                        total_scheme_stats[scheme_type] = {'total': 0, 'success': 0}
                    total_scheme_stats[scheme_type]['total'] += type_stats['total']
                    total_scheme_stats[scheme_type]['success'] += type_stats['success']
            
            if total_scheme_stats:
                scheme_data = []
                for scheme_type, stats in total_scheme_stats.items():
                    if stats['total'] > 0:
                        success_rate = round(stats['success'] / stats['total'] * 100, 2)
                    else:
                        success_rate = 0
                    
                    scheme_data.append({
                        '方案类型': scheme_type,
                        '总方案数': stats['total'],
                        '完成数': stats['success'],
                        '完成率(%)': success_rate
                    })
                
                if scheme_data:
                    scheme_df = pd.DataFrame(scheme_data)
                    st.dataframe(scheme_df)
                    
                    # 显示方案类型图表
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("##### 方案类型分布")
                        st.bar_chart(scheme_df.set_index('方案类型')['总方案数'])
                    with col2:
                        st.markdown("##### 完成率")
                        st.bar_chart(scheme_df.set_index('方案类型')['完成率(%)'])
    
    def _render_city_summary_analysis(self):
        """渲染地市汇总分析"""
        st.info("地市汇总分析功能")
        # 实现地市汇总分析逻辑
    
    def _render_grid_detail_analysis(self):
        """渲染网格明细分析"""
        st.info("网格明细分析功能")
        # 实现网格明细分析逻辑
    
    def _render_scheme_type_analysis(self):
        """渲染方案类型分析"""
        st.info("方案类型分析功能")
        # 实现方案类型分析逻辑
    
    def _render_implementation_analysis(self):
        """渲染实施结果分析"""
        st.info("实施结果分析功能")
        # 实现实施结果分析逻辑
    
    def _export_data(self, export_options):
        """导出数据（生成文件并缓存，不保存到磁盘）"""
        try:
            import tempfile
            
            # 检查数据库中是否有面板数据
            summary = self.db_manager.get_panel_data_summary()
            if not summary or summary.get('total_records', 0) == 0:
                st.error("❌ 数据库中没有面板数据，请先导入数据")
                return
            
            st.info(f"📊 数据库中共有 {summary.get('total_records', 0)} 条记录，开始生成导出文件...")
            
            # 清理旧的临时文件（如果存在）
            if 'export_file_data' in st.session_state:
                old_file_path = st.session_state['export_file_data']
                try:
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                        self.logger.info(f"已清理旧的临时文件: {old_file_path}")
                except Exception as e:
                    self.logger.warning(f"清理旧临时文件失败: {e}")
            
            # 创建导出文件
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 生成Excel文件
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认工作表
            
            # 根据选择的导出选项创建工作表
            for option in export_options:
                if option == "地市汇总评估过程得分":
                    self._create_city_summary_sheet(wb, "地市汇总评估过程得分")
                elif option == "网格明细评估过程得分":
                    self._create_grid_detail_sheet(wb, "网格明细评估过程得分")
                elif option == "方案实施统计":
                    self._create_scheme_statistics_sheet(wb, "方案实施统计")
                elif option == "网格标签统计":
                    self._create_grid_label_sheet(wb, "网格标签统计")
                elif option == "未完成的方案":
                    self._create_incomplete_scheme_sheet(wb, "未完成的方案")
            
            # 生成文件名
            filename = f"面板读取导出_{timestamp}.xlsx"
            
            # 保存到临时文件（用于缓存）
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file_path = temp_file.name
            wb.save(temp_file_path)
            temp_file.close()
            
            # 将文件路径和文件名保存到session_state
            st.session_state['export_file_data'] = temp_file_path
            st.session_state['export_filename'] = filename
            
            st.success(f"✅ 导出文件已生成，可以点击【下载导出文件】按钮下载")
            st.info(f"📄 文件名: {filename}")
            
            # 自动刷新页面以显示下载按钮
            st.rerun()
            
        except Exception as e:
            st.error(f"导出失败: {str(e)}")
            self.logger.error(f"导出失败: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _create_city_summary_sheet(self, wb, sheet_name):
        """创建地市汇总工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 获取实际数据
        data = self.db_manager.execute_query("SELECT * FROM panel_data ORDER BY city, grid_code")
        if not data:
            st.warning("数据库中没有面板数据")
            return
        
        # 获取超时方案数据
        timeout_schemes_query = """
        SELECT DISTINCT p.scheme_id, p.grid_code, p.city, t.scheme_status, t.exclude_status
        FROM panel_data p
        INNER JOIN timeout_scheme_list t ON p.scheme_id = t.scheme_id
        WHERE t.scheme_status IN ('已超时未完成', '超时已完成')
        AND (t.exclude_status IS NULL OR t.exclude_status NOT IN ('方案剔除', '剔除', '线下已完成', '方案变更'))
        """
        timeout_schemes_data = self.db_manager.execute_query(timeout_schemes_query)
        
        # 按网格分组超时方案数据
        timeout_by_grid = {}
        for timeout_record in timeout_schemes_data:
            grid_key = f"{timeout_record['city']}_{timeout_record['grid_code']}"
            if grid_key not in timeout_by_grid:
                timeout_by_grid[grid_key] = 0
            timeout_by_grid[grid_key] += 1
        
        # 按城市和网格分组数据，并区分2024年和2025年
        city_grid_data = {}
        for record in data:
            city = record['city']
            grid_code = record['grid_code']
            label = record['label']
            grid_key = f"{city}_{grid_code}"
            
            # 判断网格类型
            # 24年网格：包含24年相关标签即可（即使同时包含25年督办标签也算24年网格）
            # 包括：2024年第二批督办微网格,2025年督办微网格
            #      2024年遗留,2025年督办微网格  
            #      2024年第一批督办重要场景所在网格,2025年督办微网格
            is_2024_grid = ('2024年第二批督办微网格' in label or 
                           '2024年遗留' in label or 
                           '2024年第一批督办重要场景所在网格' in label)
            # 25年网格：只包括纯25年督办网格（不包含24年标签的）
            is_2025_grid = ('2025年督办微网格' in label and 
                          '2024年第二批督办微网格' not in label and
                          '2024年遗留' not in label and
                          '2024年第一批督办重要场景所在网格' not in label)
            
            if city not in city_grid_data:
                city_grid_data[city] = {
                    'grids_2024': {},  # 2024年网格
                    'grids_2025': {},  # 2025年网格
                    'schemes': {'规建': {'total': 0, 'success': 0}, '维护': {'total': 0, 'success': 0},
                              '天调': {'total': 0, 'success': 0}, '整治': {'total': 0, 'success': 0},
                              '优化': {'total': 0, 'success': 0}, '非无线': {'total': 0, 'success': 0}}
                }
            
            # 将网格记录添加到对应的年份组
            if is_2024_grid:
                if grid_code not in city_grid_data[city]['grids_2024']:
                    city_grid_data[city]['grids_2024'][grid_code] = {
                        'label': label,
                        'schemes': [],
                        'excluded_count': 0,
                        'offline_completed_count': 0,
                        'timeout_count': 0
                    }
                city_grid_data[city]['grids_2024'][grid_code]['schemes'].append(record)
            
            if is_2025_grid:
                if grid_code not in city_grid_data[city]['grids_2025']:
                    city_grid_data[city]['grids_2025'][grid_code] = {
                        'label': label,
                        'schemes': [],
                        'excluded_count': 0,
                        'offline_completed_count': 0,
                        'timeout_count': 0
                    }
                city_grid_data[city]['grids_2025'][grid_code]['schemes'].append(record)
            
            # 统计方案类型和剔除/线下完成数量
            # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
            scheme_category = record.get('scheme_category', '') or record.get('scheme_type', '')
            measures = record.get('measures', '') or record.get('scheme', '')
            scheme_type = self._judge_scheme_type(scheme_category, measures)
            vcisvail = record.get('vcisvail', '')
            scheme_id = record.get('scheme_id', '')
            
            # 排除包含_FATH_的方案
            if scheme_id and '_FATH_' in scheme_id:
                continue
            
            # 统计剔除方案总数（用于准确性计算）：包括剔除、方案剔除、方案变更，以及线下已完成（仅限这四类）
            if vcisvail in ['剔除', '方案剔除', '方案变更', '线下已完成']:
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    if is_2024_grid and grid_code in city_grid_data[city]['grids_2024']:
                        city_grid_data[city]['grids_2024'][grid_code]['excluded_count'] += 1
                    if is_2025_grid and grid_code in city_grid_data[city]['grids_2025']:
                        city_grid_data[city]['grids_2025'][grid_code]['excluded_count'] += 1
            
            # 统计线下完成总数
            if vcisvail == '线下已完成':
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    if is_2024_grid and grid_code in city_grid_data[city]['grids_2024']:
                        city_grid_data[city]['grids_2024'][grid_code]['offline_completed_count'] += 1
                    if is_2025_grid and grid_code in city_grid_data[city]['grids_2025']:
                        city_grid_data[city]['grids_2025'][grid_code]['offline_completed_count'] += 1
            
            # 统计方案类型（用于汇总）
            # 注意："线下已完成"的方案：所有类型都计入完成数，但只有规建、天调、维护、整治这四类才计入准确性分母
            if vcisvail not in ['剔除', '方案剔除']:
                if not scheme_id or '_FATH_' not in scheme_id:
                    if scheme_type in city_grid_data[city]['schemes']:
                        city_grid_data[city]['schemes'][scheme_type]['total'] += 1
                        # 统计完成数：所有类型的"线下已完成"都计入完成数（包括优化类、非无线类等）
                        if record['implement_results'] == '成功' or vcisvail == '线下已完成':
                            city_grid_data[city]['schemes'][scheme_type]['success'] += 1
        
        # 设置超时方案计数
        for city in city_grid_data:
            for grid_code, grid_info in city_grid_data[city]['grids_2024'].items():
                grid_key = f"{city}_{grid_code}"
                grid_info['timeout_count'] = timeout_by_grid.get(grid_key, 0)
            for grid_code, grid_info in city_grid_data[city]['grids_2025'].items():
                grid_key = f"{city}_{grid_code}"
                grid_info['timeout_count'] = timeout_by_grid.get(grid_key, 0)
        
        # 定义表头（所有部分共用相同的表头）
        headers = [
            '地市', '网格数', '过程得分', '规建类总方案数', '规建类完成数', '规建类完成率',
            '维护类总方案数', '维护类完成数', '维护类完成率', '天调类总方案数', '天调类完成数', '天调类完成率',
            '整治类总方案数', '整治类完成数', '整治类完成率', '优化类总方案数', '优化类完成数', '优化类完成率',
            '非无线类总方案数', '非无线类完成数', '非无线类完成率'
        ]
        
        # 当前行号（从1开始）
        current_row = 1
        
        # 存储每个地市的数据，用于后续生成三个部分
        city_data = {}
        
        # 为每个地市生成数据：24年网格汇总、2025年网格汇总、24+25年汇总
        for city, city_info in city_grid_data.items():
            # 统计各类型方案的实施率（总体的，用于24+25年汇总）
            scheme_rates_all = {}
            for scheme_type, type_stats in city_info['schemes'].items():
                if type_stats['total'] > 0:
                    scheme_rates_all[scheme_type] = round(type_stats['success'] / type_stats['total'] * 100, 2)
                else:
                    scheme_rates_all[scheme_type] = 0
            
            # 统计2024年网格的方案数（grids_2024包含所有有24年标签的网格，包括同时有25年标签的）
            grids_2024_only = set(city_info['grids_2024'].keys())
            schemes_2024_only = {'规建': {'total': 0, 'success': 0}, '维护': {'total': 0, 'success': 0},
                               '天调': {'total': 0, 'success': 0}, '整治': {'total': 0, 'success': 0},
                               '优化': {'total': 0, 'success': 0}, '非无线': {'total': 0, 'success': 0}}
            
            for grid_code in grids_2024_only:
                grid_info = city_info['grids_2024'][grid_code]
                for scheme in grid_info['schemes']:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
                    measures = scheme.get('measures', '') or scheme.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    vcisvail = scheme.get('vcisvail', '')
                    scheme_id = scheme.get('scheme_id', '')
                    # 排除剔除、方案剔除的方案
                    if vcisvail not in ['剔除', '方案剔除']:
                        if not scheme_id or '_FATH_' not in scheme_id:
                            if scheme_type in schemes_2024_only:
                                schemes_2024_only[scheme_type]['total'] += 1
                                # 统计完成数：所有类型的"线下已完成"都计入完成数（包括优化类、非无线类等）
                                if scheme['implement_results'] == '成功' or vcisvail == '线下已完成':
                                    schemes_2024_only[scheme_type]['success'] += 1
            
            # 统计2025年网格的方案数（包括所有有2025年标签的）
            schemes_2025_all = {'规建': {'total': 0, 'success': 0}, '维护': {'total': 0, 'success': 0},
                               '天调': {'total': 0, 'success': 0}, '整治': {'total': 0, 'success': 0},
                               '优化': {'total': 0, 'success': 0}, '非无线': {'total': 0, 'success': 0}}
            
            for grid_code, grid_info in city_info['grids_2025'].items():
                for scheme in grid_info['schemes']:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
                    measures = scheme.get('measures', '') or scheme.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    vcisvail = scheme.get('vcisvail', '')
                    scheme_id = scheme.get('scheme_id', '')
                    # 排除剔除、方案剔除的方案
                    if vcisvail not in ['剔除', '方案剔除']:
                        if not scheme_id or '_FATH_' not in scheme_id:
                            if scheme_type in schemes_2025_all:
                                schemes_2025_all[scheme_type]['total'] += 1
                                # 统计完成数：所有类型的"线下已完成"都计入完成数（包括优化类、非无线类等）
                                if scheme['implement_results'] == '成功' or vcisvail == '线下已完成':
                                    schemes_2025_all[scheme_type]['success'] += 1
            
            # 计算2024年网格的平均得分（只包含真正的2024年网格，不包括同时有25年标签的）
            scores_2024 = []
            for grid_code in grids_2024_only:
                grid_info = city_info['grids_2024'][grid_code]
                # 计算准确性分母
                schemes_2024 = {}
                for scheme in grid_info['schemes']:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
                    measures = scheme.get('measures', '') or scheme.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    vcisvail = scheme.get('vcisvail', '')
                    # 排除剔除、方案剔除的方案
                    if vcisvail not in ['剔除', '方案剔除']:
                        if scheme_type in ['规建', '维护', '天调', '整治']:
                            if scheme_type not in schemes_2024:
                                schemes_2024[scheme_type] = {'total': 0}
                            schemes_2024[scheme_type]['total'] += 1
                
                accuracy_denominator_2024 = (
                    schemes_2024.get('规建', {}).get('total', 0) +
                    schemes_2024.get('维护', {}).get('total', 0) +
                    schemes_2024.get('天调', {}).get('total', 0) +
                    schemes_2024.get('整治', {}).get('total', 0)
                )
                
                grid_score_2024, _ = self._calculate_grid_process_score(
                    grid_info['schemes'],
                    grid_info['label'],
                    grid_info['excluded_count'],
                    grid_info['offline_completed_count'],
                    grid_info['timeout_count'],
                    accuracy_denominator_2024
                )
                scores_2024.append(grid_score_2024)
            
            # 计算2025年网格的平均得分（包括所有有2025年标签的）
            scores_2025 = []
            for grid_code, grid_info in city_info['grids_2025'].items():
                # 计算准确性分母
                schemes_2025 = {}
                for scheme in grid_info['schemes']:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
                    measures = scheme.get('measures', '') or scheme.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    vcisvail = scheme.get('vcisvail', '')
                    # 排除剔除、方案剔除的方案
                    if vcisvail not in ['剔除', '方案剔除']:
                        if scheme_type in ['规建', '维护', '天调', '整治']:
                            if scheme_type not in schemes_2025:
                                schemes_2025[scheme_type] = {'total': 0}
                            schemes_2025[scheme_type]['total'] += 1
                
                accuracy_denominator_2025 = (
                    schemes_2025.get('规建', {}).get('total', 0) +
                    schemes_2025.get('维护', {}).get('total', 0) +
                    schemes_2025.get('天调', {}).get('total', 0) +
                    schemes_2025.get('整治', {}).get('total', 0)
                )
                
                grid_score_2025, _ = self._calculate_grid_process_score(
                    grid_info['schemes'],
                    grid_info['label'],
                    grid_info['excluded_count'],
                    grid_info['offline_completed_count'],
                    grid_info['timeout_count'],
                    accuracy_denominator_2025
                )
                scores_2025.append(grid_score_2025)
            
            # 计算各类型方案的实施率
            scheme_rates_2024 = {}
            for scheme_type, type_stats in schemes_2024_only.items():
                if type_stats['total'] > 0:
                    scheme_rates_2024[scheme_type] = round(type_stats['success'] / type_stats['total'] * 100, 2)
                else:
                    scheme_rates_2024[scheme_type] = 0
            
            scheme_rates_2025 = {}
            for scheme_type, type_stats in schemes_2025_all.items():
                if type_stats['total'] > 0:
                    scheme_rates_2025[scheme_type] = round(type_stats['success'] / type_stats['total'] * 100, 2)
                else:
                    scheme_rates_2025[scheme_type] = 0
            
            # 计算平均得分
            avg_score_2024 = sum(scores_2024) / len(scores_2024) if scores_2024 else 0
            avg_score_2025 = sum(scores_2025) / len(scores_2025) if scores_2025 else 0
            
            # 网格数统计
            grid_count_2024 = len(grids_2024_only)
            grid_count_2025 = len(city_info['grids_2025'])
            all_grids = set(city_info['grids_2024'].keys()) | set(city_info['grids_2025'].keys())
            grid_count_all = len(all_grids)
            
            # 计算24+25年的平均得分（所有网格的平均值）
            all_scores = []
            for grid_code in all_grids:
                # 优先使用2025年标签的网格信息，如果只有2024年标签则使用2024年信息
                if grid_code in city_info['grids_2025']:
                    grid_info = city_info['grids_2025'][grid_code]
                else:
                    grid_info = city_info['grids_2024'][grid_code]
                
                # 计算准确性分母
                schemes_all = {}
                for scheme in grid_info['schemes']:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = scheme.get('scheme_category', '') or scheme.get('scheme_type', '')
                    measures = scheme.get('measures', '') or scheme.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    vcisvail = scheme.get('vcisvail', '')
                    # 排除剔除、方案剔除的方案
                    if vcisvail not in ['剔除', '方案剔除']:
                        if scheme_type in ['规建', '维护', '天调', '整治']:
                            if scheme_type not in schemes_all:
                                schemes_all[scheme_type] = {'total': 0}
                            schemes_all[scheme_type]['total'] += 1
                
                accuracy_denominator_all = (
                    schemes_all.get('规建', {}).get('total', 0) +
                    schemes_all.get('维护', {}).get('total', 0) +
                    schemes_all.get('天调', {}).get('total', 0) +
                    schemes_all.get('整治', {}).get('total', 0)
                )
                
                grid_score_all, _ = self._calculate_grid_process_score(
                    grid_info['schemes'],
                    grid_info['label'],
                    grid_info['excluded_count'],
                    grid_info['offline_completed_count'],
                    grid_info['timeout_count'],
                    accuracy_denominator_all
                )
                all_scores.append(grid_score_all)
            
            avg_score_all = sum(all_scores) / len(all_scores) if all_scores else 0
            
            # 存储每个地市的数据，用于后续生成三个部分
            city_data[city] = {
                'grid_count_all': grid_count_all,
                'avg_score_all': avg_score_all,
                'schemes_all': city_info['schemes'],
                'scheme_rates_all': scheme_rates_all,
                'grid_count_2024': grid_count_2024,
                'avg_score_2024': avg_score_2024,
                'schemes_2024': schemes_2024_only,
                'scheme_rates_2024': scheme_rates_2024,
                'grid_count_2025': grid_count_2025,
                'avg_score_2025': avg_score_2025,
                'schemes_2025': schemes_2025_all,
                'scheme_rates_2025': scheme_rates_2025
            }
        
        # 获取排序后的地市列表（按字母顺序）
        sorted_cities = sorted(city_data.keys())
        
        # ========== 第一部分：24年+25年 ==========
        # 添加标题
        ws.cell(row=current_row, column=1, value="24年+25年")
        current_row += 1
        
        # 添加表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1
        
        # 填充24+25年数据
        for city in sorted_cities:
            city_info_data = city_data[city]
            row_data_all = [
                city,
                city_info_data['grid_count_all'],
                round(city_info_data['avg_score_all'] * 100, 2),
                city_info_data['schemes_all']['规建']['total'],
                city_info_data['schemes_all']['规建']['success'],
                city_info_data['scheme_rates_all']['规建'],
                city_info_data['schemes_all']['维护']['total'],
                city_info_data['schemes_all']['维护']['success'],
                city_info_data['scheme_rates_all']['维护'],
                city_info_data['schemes_all']['天调']['total'],
                city_info_data['schemes_all']['天调']['success'],
                city_info_data['scheme_rates_all']['天调'],
                city_info_data['schemes_all']['整治']['total'],
                city_info_data['schemes_all']['整治']['success'],
                city_info_data['scheme_rates_all']['整治'],
                city_info_data['schemes_all']['优化']['total'],
                city_info_data['schemes_all']['优化']['success'],
                city_info_data['scheme_rates_all']['优化'],
                city_info_data['schemes_all']['非无线']['total'],
                city_info_data['schemes_all']['非无线']['success'],
                city_info_data['scheme_rates_all']['非无线']
            ]
            for col, value in enumerate(row_data_all, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
        
        # 添加空行分隔
        current_row += 1
        
        # ========== 第二部分：24年 ==========
        # 添加标题
        ws.cell(row=current_row, column=1, value="24年")
        current_row += 1
        
        # 添加表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1
        
        # 填充24年数据
        for city in sorted_cities:
            city_info_data = city_data[city]
            row_data_2024 = [
                city,
                city_info_data['grid_count_2024'],
                round(city_info_data['avg_score_2024'] * 100, 2),
                city_info_data['schemes_2024']['规建']['total'],
                city_info_data['schemes_2024']['规建']['success'],
                city_info_data['scheme_rates_2024']['规建'],
                city_info_data['schemes_2024']['维护']['total'],
                city_info_data['schemes_2024']['维护']['success'],
                city_info_data['scheme_rates_2024']['维护'],
                city_info_data['schemes_2024']['天调']['total'],
                city_info_data['schemes_2024']['天调']['success'],
                city_info_data['scheme_rates_2024']['天调'],
                city_info_data['schemes_2024']['整治']['total'],
                city_info_data['schemes_2024']['整治']['success'],
                city_info_data['scheme_rates_2024']['整治'],
                city_info_data['schemes_2024']['优化']['total'],
                city_info_data['schemes_2024']['优化']['success'],
                city_info_data['scheme_rates_2024']['优化'],
                city_info_data['schemes_2024']['非无线']['total'],
                city_info_data['schemes_2024']['非无线']['success'],
                city_info_data['scheme_rates_2024']['非无线']
            ]
            for col, value in enumerate(row_data_2024, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
        
        # 添加空行分隔
        current_row += 1
        
        # ========== 第三部分：25年 ==========
        # 添加标题
        ws.cell(row=current_row, column=1, value="25年")
        current_row += 1
        
        # 添加表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1
        
        # 填充25年数据
        for city in sorted_cities:
            city_info_data = city_data[city]
            row_data_2025 = [
                city,
                city_info_data['grid_count_2025'],
                round(city_info_data['avg_score_2025'] * 100, 2),
                city_info_data['schemes_2025']['规建']['total'],
                city_info_data['schemes_2025']['规建']['success'],
                city_info_data['scheme_rates_2025']['规建'],
                city_info_data['schemes_2025']['维护']['total'],
                city_info_data['schemes_2025']['维护']['success'],
                city_info_data['scheme_rates_2025']['维护'],
                city_info_data['schemes_2025']['天调']['total'],
                city_info_data['schemes_2025']['天调']['success'],
                city_info_data['scheme_rates_2025']['天调'],
                city_info_data['schemes_2025']['整治']['total'],
                city_info_data['schemes_2025']['整治']['success'],
                city_info_data['scheme_rates_2025']['整治'],
                city_info_data['schemes_2025']['优化']['total'],
                city_info_data['schemes_2025']['优化']['success'],
                city_info_data['scheme_rates_2025']['优化'],
                city_info_data['schemes_2025']['非无线']['total'],
                city_info_data['schemes_2025']['非无线']['success'],
                city_info_data['scheme_rates_2025']['非无线']
            ]
            for col, value in enumerate(row_data_2025, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
    
    def _create_grid_detail_sheet(self, wb, sheet_name):
        """创建网格明细工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 添加表头
        headers = [
            '地市', '网格ID', '网格中文名', '网格标签', '过程得分', '规建类总方案数', '规建类完成数',
            '维护类总方案数', '维护类完成数', '天调类总方案数', '天调类完成数',
            '整治类总方案数', '整治类完成数', '优化类总方案数', '优化类完成数',
            '非无线类总方案数', '非无线类完成数', '准确性分母', '剔除方案总数', '线下完成总数', '超时方案',
            '勘察得分', '方案过程得分', '方案时效得分', '方案超时比例', '方案准确性得分', '方案准确性比例', '总权值', '备注'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 获取实际数据
        data = self.db_manager.execute_query("SELECT * FROM panel_data ORDER BY city, grid_code")
        if not data:
            st.warning("数据库中没有面板数据")
            return
        
        # 获取超时方案数据（关联timeout_scheme_list和panel_data）
        timeout_schemes_query = """
        SELECT DISTINCT p.scheme_id, p.grid_code, p.city, t.scheme_status, t.exclude_status
        FROM panel_data p
        INNER JOIN timeout_scheme_list t ON p.scheme_id = t.scheme_id
        WHERE t.scheme_status IN ('已超时未完成', '超时已完成')
        AND (t.exclude_status IS NULL OR t.exclude_status NOT IN ('方案剔除', '剔除', '线下已完成', '方案变更'))
        """
        timeout_schemes_data = self.db_manager.execute_query(timeout_schemes_query)
        
        # 按网格分组超时方案数据
        timeout_by_grid = {}
        for timeout_record in timeout_schemes_data:
            grid_key = f"{timeout_record['city']}_{timeout_record['grid_code']}"
            if grid_key not in timeout_by_grid:
                timeout_by_grid[grid_key] = 0
            timeout_by_grid[grid_key] += 1
        
        # 按网格分组数据
        grid_data = {}
        for record in data:
            grid_key = f"{record['city']}_{record['grid_code']}"
            if grid_key not in grid_data:
                grid_data[grid_key] = {
                    'city': record['city'],
                    'grid_code': record['grid_code'],
                    'grid_name': record.get('grid_name', '') or record.get('vcsmall_grid_name', '') or '',  # 网格中文名
                    'label': record['label'],
                    'schemes': {'规建': {'total': 0, 'success': 0}, '维护': {'total': 0, 'success': 0},
                              '天调': {'total': 0, 'success': 0}, '整治': {'total': 0, 'success': 0},
                              '优化': {'total': 0, 'success': 0}, '非无线': {'total': 0, 'success': 0}},
                    'excluded_count': 0,  # 剔除方案总数
                    'offline_completed_count': 0,  # 线下完成总数
                    'timeout_count': 0,  # 超时方案总数
                    'scheme_records': []  # 保存该网格的所有方案记录，用于计算过程分
                }
            
            # 保存方案记录
            grid_data[grid_key]['scheme_records'].append(record)
            
            # 判断方案类型
            # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
            scheme_category = record.get('scheme_category', '') or record.get('scheme_type', '')
            measures = record.get('measures', '') or record.get('scheme', '')
            scheme_type = self._judge_scheme_type(scheme_category, measures)
            vcisvail = record.get('vcisvail', '')
            scheme_id = record.get('scheme_id', '')
            
            # 排除包含_FATH_的方案（替换流程的工单）
            if scheme_id and '_FATH_' in scheme_id:
                continue
            
            # 统计剔除方案总数（规建类、维护类、天调类、整治类）
            # 用于准确性计算：包括剔除、方案剔除、方案变更，以及线下已完成（仅限这四类）
            if vcisvail in ['剔除', '方案剔除', '方案变更', '线下已完成']:
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    grid_data[grid_key]['excluded_count'] += 1
            
            # 统计线下完成总数（规建类、维护类、天调类、整治类）
            # vcisvail = 线下已完成
            if vcisvail == '线下已完成':
                if scheme_type in ['规建', '维护', '天调', '整治']:
                    grid_data[grid_key]['offline_completed_count'] += 1
            
            # 判断方案类型 - 排除剔除、方案剔除的数据，以及包含_FATH_的方案
            # 注意："线下已完成"的方案：所有类型都计入完成数，但只有规建、天调、维护、整治这四类才计入准确性分母
            if vcisvail not in ['剔除', '方案剔除']:
                # 排除包含_FATH_的方案（替换流程的工单）
                if not scheme_id or '_FATH_' not in scheme_id:
                    if scheme_type in grid_data[grid_key]['schemes']:
                        grid_data[grid_key]['schemes'][scheme_type]['total'] += 1
                        # 统计完成数：所有类型的"线下已完成"都计入完成数（包括优化类、非无线类等）
                        if record['implement_results'] == '成功' or vcisvail == '线下已完成':
                            grid_data[grid_key]['schemes'][scheme_type]['success'] += 1
        
        # 设置超时方案计数
        for grid_key in grid_data:
            grid_data[grid_key]['timeout_count'] = timeout_by_grid.get(grid_key, 0)
        
        # 定义样式：黄色背景 + 红色字体（用于未满分的单元格）
        warning_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        warning_font = Font(color='FF0000', bold=False)
        
        # 填充数据
        row = 2
        for grid_key, grid_info in grid_data.items():
            # 计算准确性分母（规建类、维护类、天调类、整治类总方案数相加）
            # 注意：准确性分母已经包含了"线下已完成"的方案（在统计total时已包含）
            accuracy_denominator = (
                grid_info['schemes']['规建']['total'] +
                grid_info['schemes']['维护']['total'] +
                grid_info['schemes']['天调']['total'] +
                grid_info['schemes']['整治']['total']
            )
            
            # 使用新公式计算过程得分
            process_score, score_details = self._calculate_grid_process_score(
                grid_info['scheme_records'],
                grid_info['label'],
                grid_info['excluded_count'],
                grid_info['offline_completed_count'],
                grid_info['timeout_count'],
                accuracy_denominator
            )
            
            # 判断网格类型，24年网格不考核方案时效得分和方案准确性得分
            is_2025_grid = score_details.get('is_2025_grid', False)
            
            # 填充行数据
            scheme_process_score_val = round(score_details.get('scheme_process_score', 0), 2)
            scheme_process_weight_val = score_details.get('scheme_process_weight', 0)  # 方案过程分实际权重
            timeliness_score_val = round(score_details.get('timeliness_score', 0), 2) if is_2025_grid else ''
            accuracy_score_val = round(score_details.get('accuracy_score', 0), 2) if is_2025_grid else ''
            
            # 计算方案超时比例（超时方案数 / 总方案数）
            # 总方案数 = 所有类型方案数之和（排除剔除和方案剔除的）
            total_schemes_count = (
                grid_info['schemes']['规建']['total'] +
                grid_info['schemes']['维护']['total'] +
                grid_info['schemes']['天调']['total'] +
                grid_info['schemes']['整治']['total'] +
                grid_info['schemes']['优化']['total'] +
                grid_info['schemes']['非无线']['total']
            )
            timeout_ratio = round(grid_info['timeout_count'] / total_schemes_count, 4) if total_schemes_count > 0 else 0.0
            timeout_ratio_val = timeout_ratio if is_2025_grid else ''  # 24年网格不显示
            
            # 计算方案准确性比例（剔除方案总数 / (准确性分母 + 剔除方案总数)）
            accuracy_total = accuracy_denominator + grid_info['excluded_count']
            accuracy_ratio = round(grid_info['excluded_count'] / accuracy_total, 4) if accuracy_total > 0 else 0.0
            accuracy_ratio_val = accuracy_ratio if is_2025_grid else ''  # 24年网格不显示
            
            # 计算扣分项并生成备注
            deduction_items = []
            
            # 1. 方案过程得分扣分
            if scheme_process_weight_val > 0 and scheme_process_score_val < scheme_process_weight_val:
                deduction = round(scheme_process_weight_val - scheme_process_score_val, 2)
                deduction_items.append(f"过程得分失{deduction}分")
            
            # 2. 方案时效得分扣分（仅25年网格）
            if is_2025_grid and isinstance(timeliness_score_val, (int, float)) and timeliness_score_val < 10:
                deduction = round(10 - timeliness_score_val, 2)
                deduction_items.append(f"方案时效性失{deduction}分")
            
            # 3. 方案准确性得分扣分（仅25年网格）
            if is_2025_grid and isinstance(accuracy_score_val, (int, float)) and accuracy_score_val < 10:
                deduction = round(10 - accuracy_score_val, 2)
                deduction_items.append(f"方案准确性失{deduction}分")
            
            # 生成备注文字
            if deduction_items:
                if len(deduction_items) == 1:
                    remark = deduction_items[0]
                else:
                    remark = "多项，" + "，".join(deduction_items)
            else:
                remark = ''  # 无扣分项时备注为空
            
            row_data = [
                grid_info['city'], grid_info['grid_code'], grid_info['grid_name'], grid_info['label'], round(process_score * 100, 2),
                grid_info['schemes']['规建']['total'], grid_info['schemes']['规建']['success'],
                grid_info['schemes']['维护']['total'], grid_info['schemes']['维护']['success'],
                grid_info['schemes']['天调']['total'], grid_info['schemes']['天调']['success'],
                grid_info['schemes']['整治']['total'], grid_info['schemes']['整治']['success'],
                grid_info['schemes']['优化']['total'], grid_info['schemes']['优化']['success'],
                grid_info['schemes']['非无线']['total'], grid_info['schemes']['非无线']['success'],
                accuracy_denominator,  # 准确性分母
                grid_info['excluded_count'],  # 剔除方案总数
                grid_info['offline_completed_count'],  # 线下完成总数
                grid_info['timeout_count'],  # 超时方案
                round(score_details.get('kc_score', 0), 2),  # 勘察得分
                scheme_process_score_val,  # 方案过程得分
                timeliness_score_val,  # 方案时效得分（24年网格留空）
                timeout_ratio_val,  # 方案超时比例（24年网格留空）
                accuracy_score_val,  # 方案准确性得分（24年网格留空）
                accuracy_ratio_val,  # 方案准确性比例（24年网格留空）
                round(score_details.get('total_weight', 0), 2),  # 总权值
                remark  # 备注
            ]
            
            # 填充数据并设置样式
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # 判断是否需要设置警告样式
                # 方案过程得分（第23列）：基于实际权重判断（列索引因增加了网格中文名而+1）
                if col == 23 and isinstance(value, (int, float)) and scheme_process_weight_val > 0:
                    if value < scheme_process_weight_val:  # 得分小于实际权重，说明未满分
                        cell.fill = warning_fill
                        cell.font = warning_font
                
                # 方案时效得分（第24列）：满分10分（仅25年网格）
                elif col == 24 and isinstance(value, (int, float)) and value < 10:
                    cell.fill = warning_fill
                    cell.font = warning_font
                
                # 方案准确性得分（第26列）：满分10分（仅25年网格）
                elif col == 26 and isinstance(value, (int, float)) and value < 10:
                    cell.fill = warning_fill
                    cell.font = warning_font
            
            row += 1
    
    def _create_incomplete_scheme_sheet(self, wb, sheet_name):
        """创建未完成的方案工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 添加表头
        headers = [
            '地市', '网格ID', '网格中文名', '方案ID', '方案类型', '优化对象名称', '方案', '方案状态',
            '实施结果', '方案标签', 'vcmeasure_code', '提交时间', '是否超时'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 查询未完成的方案
        # 条件：implement_results为空或失败，且vcisvail不在【剔除、方案剔除、方案变更】中
        # 注意："线下已完成"的方案应该算作已完成，不应该出现在未完成列表中
        # 同时关联超时方案表判断是否超时
        incomplete_schemes_query = """
        SELECT 
            p.city, p.grid_code, 
            COALESCE(p.grid_name, '') as grid_name,
            p.scheme_id, p.scheme_type, p.vcoptimize_object_name, p.scheme,
            p.vcisvail, p.implement_results, p.label, p.vcmeasure_code, p.created_at,
            CASE 
                WHEN t.scheme_id IS NOT NULL 
                    AND t.scheme_status IN ('已超时未完成', '超时已完成')
                    AND (t.exclude_status IS NULL OR t.exclude_status NOT IN ('方案剔除', '剔除', '方案变更'))
                THEN '是'
                ELSE '否'
            END as is_timeout
        FROM panel_data p
        LEFT JOIN timeout_scheme_list t ON p.scheme_id = t.scheme_id
        WHERE (
            p.implement_results IS NULL 
            OR p.implement_results = '' 
            OR p.implement_results = 'NULL'
            OR p.implement_results = '失败'
            OR LOWER(TRIM(p.implement_results)) = '失败'
        )
        AND (p.vcisvail IS NULL OR p.vcisvail NOT IN ('剔除', '方案剔除', '方案变更', '线下已完成'))
        AND (p.scheme_id IS NULL OR p.scheme_id NOT LIKE '%_FATH_%')
        ORDER BY p.city, p.grid_code, p.scheme_id
        """
        
        incomplete_schemes = self.db_manager.execute_query(incomplete_schemes_query)
        
        if not incomplete_schemes:
            ws.cell(row=2, column=1, value="暂无未完成的方案")
            return
        
        # 填充数据
        row = 2
        for scheme in incomplete_schemes:
            row_data = [
                scheme.get('city', ''),
                scheme.get('grid_code', ''),
                scheme.get('grid_name', '') or scheme.get('vcsmall_grid_name', ''),
                scheme.get('scheme_id', ''),
                scheme.get('scheme_type', ''),
                scheme.get('vcoptimize_object_name', '') or '',  # 优化对象名称
                scheme.get('scheme', ''),
                scheme.get('vcisvail', ''),
                scheme.get('implement_results', ''),
                scheme.get('label', ''),
                scheme.get('vcmeasure_code', '') or '',  # vcmeasure_code
                scheme.get('created_at', ''),  # 使用created_at代替scheme_submit_time
                scheme.get('is_timeout', '否')  # 是否超时
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
    
    def _create_scheme_statistics_sheet(self, wb, sheet_name):
        """创建方案统计工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 添加表头
        headers = ['方案类型', '总方案数', '完成数', '完成率(%)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 获取实际数据
        data = self.db_manager.execute_query("SELECT * FROM panel_data ORDER BY city, grid_code")
        if not data:
            st.warning("数据库中没有面板数据")
            return
        
        # 统计各类型方案
        scheme_stats = {'规建': {'total': 0, 'success': 0}, '维护': {'total': 0, 'success': 0},
                      '天调': {'total': 0, 'success': 0}, '整治': {'total': 0, 'success': 0},
                      '优化': {'total': 0, 'success': 0}, '非无线': {'total': 0, 'success': 0}}
        
        for record in data:
            # 排除剔除、方案剔除的数据，以及包含_FATH_的方案
            # 注意："线下已完成"的方案：所有类型都计入完成数，但只有规建、天调、维护、整治这四类才计入准确性分母
            if record.get('vcisvail') not in ['剔除', '方案剔除']:
                scheme_id = record.get('scheme_id', '')
                # 排除包含_FATH_的方案（替换流程的工单）
                if scheme_id and '_FATH_' not in scheme_id:
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = record.get('scheme_category', '') or record.get('scheme_type', '')
                    measures = record.get('measures', '') or record.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    if scheme_type in scheme_stats:
                        scheme_stats[scheme_type]['total'] += 1
                        # 统计完成数：所有类型的"线下已完成"都计入完成数（包括优化类、非无线类等）
                        vcisvail = record.get('vcisvail', '')
                        if record['implement_results'] == '成功' or vcisvail == '线下已完成':
                            scheme_stats[scheme_type]['success'] += 1
        
        # 填充数据
        row = 2
        for scheme_type, stats in scheme_stats.items():
            if stats['total'] > 0:
                success_rate = round(stats['success'] / stats['total'] * 100, 2)
            else:
                success_rate = 0
            
            row_data = [scheme_type, stats['total'], stats['success'], success_rate]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
    
    def _create_grid_label_sheet(self, wb, sheet_name):
        """创建网格标签工作表"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 添加表头
        headers = ['网格标签', '网格数', '平均得分', '总方案数', '完成方案数', '完成率(%)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 获取实际数据
        data = self.db_manager.execute_query("SELECT * FROM panel_data ORDER BY city, grid_code")
        if not data:
            st.warning("数据库中没有面板数据")
            return
        
        # 按标签分组数据
        label_data = {}
        for record in data:
            label = record['label']
            if label not in label_data:
                label_data[label] = {
                    'grids': set(),
                    'schemes': {'total': 0, 'success': 0},
                    'scores': []
                }
            
            label_data[label]['grids'].add(f"{record['city']}_{record['grid_code']}")
            
            # 排除剔除、方案剔除的数据，以及包含_FATH_的方案，不参与统计和得分计算
            # 注意："线下已完成"的方案：所有类型都计入完成数，但只有规建、天调、维护、整治这四类才计入准确性分母
            if record.get('vcisvail') not in ['剔除', '方案剔除']:
                scheme_id = record.get('scheme_id', '')
                # 排除包含_FATH_的方案（替换流程的工单）
                if scheme_id and '_FATH_' not in scheme_id:
                    label_data[label]['schemes']['total'] += 1
                    # 统计完成数：所有类型的"线下已完成"都计入完成数（包括优化类、非无线类等）
                    vcisvail = record.get('vcisvail', '')
                    # 使用scheme_category（方案分类）和measures（措施）来判断方案类型
                    scheme_category = record.get('scheme_category', '') or record.get('scheme_type', '')
                    measures = record.get('measures', '') or record.get('scheme', '')
                    scheme_type = self._judge_scheme_type(scheme_category, measures)
                    
                    # 所有类型：统计"成功"和"线下已完成"
                    if record['implement_results'] == '成功' or vcisvail == '线下已完成':
                        label_data[label]['schemes']['success'] += 1
                    
                    # 计算得分（简化版）：只有规建、天调、维护、整治这四类参与得分计算
                    if scheme_type in ['规建', '天调', '维护', '整治']:
                        score = 1.0 if (record['implement_results'] == '成功' or vcisvail == '线下已完成') else 0.0
                        label_data[label]['scores'].append(score)
        
        # 填充数据
        row = 2
        for label, stats in label_data.items():
            grid_count = len(stats['grids'])
            avg_score = round(sum(stats['scores']) / len(stats['scores']) * 100, 2) if stats['scores'] else 0
            total_schemes = stats['schemes']['total']
            success_schemes = stats['schemes']['success']
            success_rate = round(success_schemes / total_schemes * 100, 2) if total_schemes > 0 else 0
            
            row_data = [label, grid_count, avg_score, total_schemes, success_schemes, success_rate]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
    
    def _render_data_query(self):
        """渲染数据查询页面"""
        st.subheader("🔍 数据查询")
        
        # 查询类型选择
        query_type = st.radio(
            "选择查询类型",
            ["基础查询", "高级查询", "统计查询", "批次管理"],
            horizontal=True
        )
        
        if query_type == "基础查询":
            self._render_basic_query()
        elif query_type == "高级查询":
            self._render_advanced_query()
        elif query_type == "统计查询":
            self._render_statistics_query()
        elif query_type == "批次管理":
            self._render_batch_management()
    
    def _render_basic_query(self):
        """渲染基础查询页面"""
        st.markdown("#### 基础查询")
        
        # 查询条件
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # 地市选择
            cities = self._get_available_cities()
            selected_city = st.selectbox("选择地市", ["全部"] + cities)
            
            # 网格代码
            grid_code = st.text_input("网格代码", placeholder="输入网格代码进行精确查询")
            
            # 方案类型
            scheme_types = self._get_available_scheme_types()
            selected_scheme_type = st.selectbox("方案类型", ["全部"] + scheme_types)
            
            # 实施结果
            implement_results = self._get_available_implement_results()
            selected_implement_result = st.selectbox("实施结果", ["全部", "空值"] + implement_results)
        
        with col2:
            # 标签模式
            label_pattern = st.text_input("标签模式", placeholder="输入标签关键词，支持模糊查询")
            
            # 批次选择
            batches = self._get_available_batches()
            selected_batch = st.selectbox("选择批次", ["最新批次"] + [batch['batch_id'] for batch in batches])
            
            # 优化对象名称
            vcoptimize_object_names = self._get_available_vcoptimize_object_names()
            selected_vcoptimize_object_name = st.selectbox("优化对象名称", ["全部"] + vcoptimize_object_names)
            
            # vcisvail
            vcisvail_values = self._get_available_vcisvail_values()
            selected_vcisvail = st.selectbox("vcisvail", ["全部", "空值"] + vcisvail_values)
        
        with col3:
            # vcmeasure_code
            vcmeasure_codes = self._get_available_vcmeasure_codes()
            selected_vcmeasure_code = st.selectbox("vcmeasure_code", ["全部"] + vcmeasure_codes)
            
            # current_act_name
            current_act_names = self._get_available_current_act_names()
            selected_current_act_name = st.selectbox("current_act_name", ["全部"] + current_act_names)
        
        # 查询按钮
        if st.button("🔍 执行查询", type="primary"):
            self._execute_basic_query({
                'city': selected_city if selected_city != "全部" else None,
                'grid_code': grid_code if grid_code else None,
                'scheme_type': selected_scheme_type if selected_scheme_type != "全部" else None,
                'implement_results': selected_implement_result if selected_implement_result not in ["全部", "空值"] else None,
                'implement_results_is_null': selected_implement_result == "空值",
                'label_pattern': label_pattern if label_pattern else None,
                'batch_id': selected_batch if selected_batch != "最新批次" else None,
                'vcoptimize_object_name': selected_vcoptimize_object_name if selected_vcoptimize_object_name != "全部" else None,
                'vcisvail': selected_vcisvail if selected_vcisvail not in ["全部", "空值"] else None,
                'vcisvail_is_null': selected_vcisvail == "空值",
                'vcmeasure_code': selected_vcmeasure_code if selected_vcmeasure_code != "全部" else None,
                'current_act_name': selected_current_act_name if selected_current_act_name != "全部" else None
            })
    
    def _render_advanced_query(self):
        """渲染高级查询页面"""
        st.markdown("#### 高级查询")
        
        # 自定义SQL查询
        st.markdown("##### 自定义SQL查询")
        st.info("⚠️ 高级功能：请谨慎使用SQL查询，确保查询语句正确")
        
        # 预设查询模板
        query_template = st.selectbox(
            "选择查询模板",
            ["自定义", "按地市统计", "按网格统计", "按方案类型统计", "按实施结果统计"]
        )
        
        if query_template == "自定义":
            sql_query = st.text_area(
                "SQL查询语句",
                placeholder="SELECT * FROM panel_data WHERE city = '北京' LIMIT 100",
                height=100
            )
        else:
            sql_query = self._get_query_template(query_template)
            st.text_area("SQL查询语句", value=sql_query, height=100)
        
        if st.button("🚀 执行SQL查询", type="primary"):
            if sql_query.strip():
                self._execute_sql_query(sql_query)
            else:
                st.error("请输入SQL查询语句")
    
    def _render_statistics_query(self):
        """渲染统计查询页面"""
        st.markdown("#### 统计查询")
        
        # 统计类型选择
        stat_type = st.radio(
            "选择统计类型",
            ["地市统计", "网格统计", "方案类型统计", "实施结果统计", "时间趋势统计"],
            horizontal=True
        )
        
        if stat_type == "地市统计":
            self._show_city_statistics()
        elif stat_type == "网格统计":
            self._show_grid_statistics()
        elif stat_type == "方案类型统计":
            self._show_scheme_type_statistics()
        elif stat_type == "实施结果统计":
            self._show_implement_result_statistics()
        elif stat_type == "时间趋势统计":
            self._show_time_trend_statistics()
    
    def _render_batch_management(self):
        """渲染批次管理页面"""
        st.markdown("#### 批次管理")
        
        # 显示批次列表
        batches = self._get_available_batches()
        
        if batches:
            st.markdown("##### 导入批次列表")
            batch_df = pd.DataFrame(batches)
            st.dataframe(batch_df, use_container_width=True)
            
            # 批次操作
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("🔄 刷新批次列表"):
                    st.rerun()
            
            with col2:
                if st.button("🗑️ 清空所有数据", type="secondary"):
                    if st.session_state.get('confirm_clear', False):
                        if self.db_manager.clear_panel_data():
                            st.success("✅ 所有数据已清空")
                            st.session_state['confirm_clear'] = False
                        else:
                            st.error("❌ 清空数据失败")
                    else:
                        st.session_state['confirm_clear'] = True
                        st.warning("⚠️ 点击确认清空所有数据")
            
            with col3:
                if st.button("📊 查看数据库统计"):
                    self._show_database_statistics()
        else:
            st.info("📭 暂无导入批次数据")
    
    def _get_available_cities(self):
        """获取可用的地市列表"""
        try:
            sql = "SELECT DISTINCT city FROM panel_data ORDER BY city"
            results = self.db_manager.execute_query(sql)
            return [row['city'] for row in results]
        except:
            return []
    
    def _get_available_scheme_types(self):
        """获取可用的方案类型列表"""
        try:
            sql = "SELECT DISTINCT scheme_type FROM panel_data WHERE scheme_type IS NOT NULL ORDER BY scheme_type"
            results = self.db_manager.execute_query(sql)
            return [row['scheme_type'] for row in results]
        except:
            return []
    
    def _get_available_implement_results(self):
        """获取可用的实施结果列表"""
        try:
            sql = "SELECT DISTINCT implement_results FROM panel_data WHERE implement_results IS NOT NULL ORDER BY implement_results"
            results = self.db_manager.execute_query(sql)
            return [row['implement_results'] for row in results]
        except:
            return []
    
    def _get_available_batches(self):
        """获取可用的批次列表"""
        try:
            return self.db_manager.get_panel_import_batches(20)
        except:
            return []
    
    def _get_available_vcoptimize_object_names(self):
        """获取可用的优化对象名称列表"""
        try:
            return self.db_manager.get_available_vcoptimize_object_names()
        except:
            return []
    
    def _get_available_vcisvail_values(self):
        """获取可用的vcisvail值列表"""
        try:
            return self.db_manager.get_available_vcisvail_values()
        except:
            return []
    
    def _get_available_vcmeasure_codes(self):
        """获取可用的vcmeasure_code列表"""
        try:
            return self.db_manager.get_available_vcmeasure_codes()
        except:
            return []
    
    def _get_available_current_act_names(self):
        """获取可用的current_act_name列表"""
        try:
            return self.db_manager.get_available_current_act_names()
        except:
            return []
    
    def _execute_basic_query(self, filters):
        """执行基础查询"""
        try:
            # 获取批次ID
            batch_id = None
            if filters.get('batch_id'):
                batch_id = filters['batch_id']
            elif 'current_batch_id' in st.session_state:
                batch_id = st.session_state['current_batch_id']
            
            # 执行查询
            results = self.db_manager.search_panel_data(filters, batch_id)
            
            if results:
                st.success(f"✅ 查询完成，共找到 {len(results)} 条记录")
                
                # 显示结果
                df = pd.DataFrame(results)
                
                # 定义中文列名映射
                column_mapping = {
                    'id': 'ID',
                    'city': '地市',
                    'grid_code': '网格代码',
                    'grid_name': '网格名称',
                    'scheme_type': '方案类型',
                    'scheme': '方案',
                    'order_type': '工单类型',
                    'order_status': '工单状态',
                    'implement_results': '实施结果',
                    'scheme_id': '方案ID',
                    'is_valid': '是否有效',
                    'label': '标签',
                    'filename': '文件名',
                    'import_batch_id': '导入批次ID',
                    'vcoptimize_object_name': '优化对象名称',
                    'vcisvail': 'vcisvail',
                    'vcmeasure_code': 'vcmeasure_code',
                    'current_act_name': '当前活动名称',
                    'created_at': '创建时间',
                    'updated_at': '更新时间',
                    'scheme_submit_time': '方案提交时间',
                    'scheme_complete_time': '方案完成时间',
                    'is_timeout': '是否超时'
                }
                
                # 重命名列为中文
                df = df.rename(columns=column_mapping)
                
                # 重新排列列的顺序，将重要列放在前面
                priority_cols = ['地市', '网格代码', '网格名称', '方案类型', '方案', '优化对象名称', 
                                '方案提交时间', '方案完成时间', '是否超时']
                existing_priority_cols = [col for col in priority_cols if col in df.columns]
                other_cols = [col for col in df.columns if col not in existing_priority_cols]
                df = df[existing_priority_cols + other_cols]
                
                # 应用样式
                def highlight_cells(row):
                    styles = [''] * len(row)
                    
                    # 处理"是否超时"列：值为"是"时字体变红
                    if '是否超时' in df.columns:
                        timeout_idx = df.columns.get_loc('是否超时')
                        if row['是否超时'] == '是':
                            styles[timeout_idx] = 'color: red; font-weight: bold'
                    
                    # 处理"是否有效"列
                    if '是否有效' in df.columns:
                        valid_idx = df.columns.get_loc('是否有效')
                        valid_value = str(row['是否有效']).strip()
                        
                        # "剔除相关的"字体变红
                        if valid_value == '剔除相关的':
                            styles[valid_idx] = 'color: red; font-weight: bold'
                        # "线下完成的"背景变绿
                        elif valid_value == '线下完成的':
                            styles[valid_idx] = 'background-color: lightgreen'
                    
                    return styles
                
                # 检查数据量，决定是否应用样式
                total_cells = len(df) * len(df.columns)
                max_cells_for_styling = 262144  # pandas styler默认限制
                
                if total_cells > max_cells_for_styling:
                    # 数据量过大，不应用样式直接显示
                    st.warning(f"⚠️ 数据量较大（{total_cells:,} 个单元格），为提升性能已取消样式渲染")
                    st.dataframe(df, use_container_width=True)
                else:
                    # 数据量适中，应用样式显示
                    styled_df = df.style.apply(highlight_cells, axis=1)
                    st.dataframe(styled_df, use_container_width=True)
                
                # 提供下载功能
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 下载查询结果",
                    data=csv,
                    file_name=f"面板数据查询结果_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
            else:
                st.warning("⚠️ 未找到符合条件的记录")
                
        except Exception as e:
            st.error(f"查询失败: {str(e)}")
            self.logger.error(f"基础查询失败: {str(e)}")
    
    def _execute_sql_query(self, sql_query):
        """执行SQL查询"""
        try:
            results = self.db_manager.execute_query(sql_query)
            
            if results:
                st.success(f"✅ 查询完成，共找到 {len(results)} 条记录")
                
                # 显示结果
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
                
                # 提供下载功能
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 下载查询结果",
                    data=csv,
                    file_name=f"SQL查询结果_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
            else:
                st.warning("⚠️ 查询未返回结果")
                
        except Exception as e:
            st.error(f"SQL查询失败: {str(e)}")
            self.logger.error(f"SQL查询失败: {str(e)}")
    
    def _get_query_template(self, template_name):
        """获取查询模板"""
        templates = {
            "按地市统计": """
SELECT 
    city as 地市,
    COUNT(*) as 记录数,
    COUNT(DISTINCT grid_code) as 网格数,
    COUNT(DISTINCT scheme_type) as 方案类型数
FROM panel_data 
GROUP BY city 
ORDER BY 记录数 DESC
            """,
            "按网格统计": """
SELECT 
    grid_code as 网格代码,
    grid_name as 网格名称,
    city as 地市,
    COUNT(*) as 方案数,
    COUNT(CASE WHEN implement_results = '成功' THEN 1 END) as 成功数
FROM panel_data 
GROUP BY grid_code, grid_name, city 
ORDER BY 方案数 DESC
LIMIT 50
            """,
            "按方案类型统计": """
SELECT 
    scheme_type as 方案类型,
    COUNT(*) as 总数,
    COUNT(CASE WHEN implement_results = '成功' THEN 1 END) as 成功数,
    ROUND(COUNT(CASE WHEN implement_results = '成功' THEN 1 END) * 100.0 / COUNT(*), 2) as 成功率
FROM panel_data 
WHERE scheme_type IS NOT NULL
GROUP BY scheme_type 
ORDER BY 总数 DESC
            """,
            "按实施结果统计": """
SELECT 
    implement_results as 实施结果,
    COUNT(*) as 数量,
    ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM panel_data), 2) as 占比
FROM panel_data 
WHERE implement_results IS NOT NULL
GROUP BY implement_results 
ORDER BY 数量 DESC
            """
        }
        return templates.get(template_name, "")
    
    def _show_city_statistics(self):
        """显示地市统计"""
        try:
            sql = """
            SELECT 
                city as 地市,
                COUNT(*) as 总记录数,
                COUNT(DISTINCT grid_code) as 网格数,
                COUNT(DISTINCT scheme_type) as 方案类型数,
                COUNT(CASE WHEN implement_results = '成功' THEN 1 END) as 成功方案数,
                ROUND(COUNT(CASE WHEN implement_results = '成功' THEN 1 END) * 100.0 / COUNT(*), 2) as 成功率
            FROM panel_data 
            GROUP BY city 
            ORDER BY 总记录数 DESC
            """
            results = self.db_manager.execute_query(sql)
            
            if results:
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
                
                # 显示图表
                col1, col2 = st.columns(2)
                with col1:
                    st.bar_chart(df.set_index('地市')['总记录数'])
                with col2:
                    st.bar_chart(df.set_index('地市')['成功率'])
            else:
                st.warning("暂无数据")
        except Exception as e:
            st.error(f"地市统计查询失败: {str(e)}")
    
    def _show_grid_statistics(self):
        """显示网格统计"""
        try:
            sql = """
            SELECT 
                grid_code as 网格代码,
                grid_name as 网格名称,
                city as 地市,
                COUNT(*) as 方案数,
                COUNT(CASE WHEN implement_results = '成功' THEN 1 END) as 成功数,
                ROUND(COUNT(CASE WHEN implement_results = '成功' THEN 1 END) * 100.0 / COUNT(*), 2) as 成功率
            FROM panel_data 
            GROUP BY grid_code, grid_name, city 
            ORDER BY 方案数 DESC
            LIMIT 100
            """
            results = self.db_manager.execute_query(sql)
            
            if results:
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
            else:
                st.warning("暂无数据")
        except Exception as e:
            st.error(f"网格统计查询失败: {str(e)}")
    
    def _show_scheme_type_statistics(self):
        """显示方案类型统计"""
        try:
            sql = """
            SELECT 
                scheme_type as 方案类型,
                COUNT(*) as 总数,
                COUNT(CASE WHEN implement_results = '成功' THEN 1 END) as 成功数,
                ROUND(COUNT(CASE WHEN implement_results = '成功' THEN 1 END) * 100.0 / COUNT(*), 2) as 成功率
            FROM panel_data 
            WHERE scheme_type IS NOT NULL
            GROUP BY scheme_type 
            ORDER BY 总数 DESC
            """
            results = self.db_manager.execute_query(sql)
            
            if results:
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
                
                # 显示饼图
                st.pie_chart(df.set_index('方案类型')['总数'])
            else:
                st.warning("暂无数据")
        except Exception as e:
            st.error(f"方案类型统计查询失败: {str(e)}")
    
    def _show_implement_result_statistics(self):
        """显示实施结果统计"""
        try:
            sql = """
            SELECT 
                implement_results as 实施结果,
                COUNT(*) as 数量,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM panel_data), 2) as 占比
            FROM panel_data 
            WHERE implement_results IS NOT NULL
            GROUP BY implement_results 
            ORDER BY 数量 DESC
            """
            results = self.db_manager.execute_query(sql)
            
            if results:
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
                
                # 显示饼图
                st.pie_chart(df.set_index('实施结果')['数量'])
            else:
                st.warning("暂无数据")
        except Exception as e:
            st.error(f"实施结果统计查询失败: {str(e)}")
    
    def _show_time_trend_statistics(self):
        """显示时间趋势统计"""
        try:
            sql = """
            SELECT 
                DATE(created_at) as 日期,
                COUNT(*) as 记录数
            FROM panel_data 
            GROUP BY DATE(created_at)
            ORDER BY 日期 DESC
            LIMIT 30
            """
            results = self.db_manager.execute_query(sql)
            
            if results:
                df = pd.DataFrame(results)
                st.dataframe(df, use_container_width=True)
                
                # 显示趋势图
                st.line_chart(df.set_index('日期')['记录数'])
            else:
                st.warning("暂无数据")
        except Exception as e:
            st.error(f"时间趋势统计查询失败: {str(e)}")
    
    def _show_database_statistics(self):
        """显示数据库统计信息"""
        try:
            stats = self.db_manager.get_database_stats()
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("面板数据记录数", stats.get('panel_data_count', 0))
            with col2:
                st.metric("导入批次数", stats.get('panel_import_batches_count', 0))
            with col3:
                st.metric("评估结果数", stats.get('panel_evaluation_results_count', 0))
            with col4:
                st.metric("数据库大小(MB)", f"{stats.get('db_size_mb', 0):.2f}")
                
        except Exception as e:
            st.error(f"获取数据库统计失败: {str(e)}")
