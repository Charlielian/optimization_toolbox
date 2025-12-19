# -*- coding: utf-8 -*-
"""
工参整理工具 - 整合到优化百宝箱工具集
提供工参数据的整理、查询和导出功能
"""

import logging
import os
import tempfile
from datetime import datetime
from typing import List
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class GCOrganizer:
    """工参整理工具类"""

    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.logger = logging.getLogger(__name__)
        self.gc_dict = {}

    def render(self):
        """渲染工参集合界面"""
        st.title("📋 工参集合")
        st.caption("工参数据查询和导出（数据导入请使用总控面板统一导入功能）")

        # 显示提示信息
        st.info("💡 **重要提示**：工参数据导入功能已移至总控面板，请到【📊 总控面板】→【⚙️ 工参表】进行数据导入")

        # 创建选项卡
        tab1, tab2, tab3 = st.tabs(["🔍 站点查询", "📊 批量导出", "📋 SQL模板导出"])

        with tab1:
            self._render_site_query()

        with tab2:
            self._render_batch_export()
        
        with tab3:
            self._render_sql_export()


    def _render_site_query(self):
        """渲染站点查询界面"""
        st.markdown("### 🔍 站点查询")

        # 从数据库获取工参数据
        try:
            # 获取所有工参数据
            engineering_data = self.db_manager.execute_query(
                "SELECT * FROM engineering_params ORDER BY phy_name, cgi"
            )
            
            if not engineering_data:
                st.warning("⚠️ 数据库中暂无工参数据，请先在总控面板导入工参数据")
                return

            # 获取所有站点名称
            site_names = list(set([row['phy_name'] for row in engineering_data if row['phy_name']]))
            site_names.sort()

            st.info(f"📊 数据库中共有 {len(engineering_data)} 条工参记录，{len(site_names)} 个站点")

            # 搜索框
            search_query = st.text_input(
                "输入站点名称进行搜索",
                placeholder="支持模糊搜索...",
                key="gc_search"
            )

            if search_query:
                # 执行搜索
                matches = self._smart_search(search_query, site_names)

                if not matches:
                    st.warning(f"未找到包含 '{search_query}' 的站点")
                    return

                # 显示搜索结果
                st.info(f"找到 {len(matches)} 个匹配的站点")

                # 选择站点
                selected_site = st.selectbox(
                    "选择站点",
                    options=matches,
                    key="selected_site"
                )

                if selected_site:
                    # 显示站点信息
                    self._display_site_info_from_db(selected_site, engineering_data)
            else:
                # 显示所有站点列表
                st.markdown("#### 📋 所有站点列表")
                if st.button("显示所有站点", key="show_all_sites"):
                    # 分页显示站点
                    page_size = 50
                    total_pages = (len(site_names) + page_size - 1) // page_size
                    
                    page = st.selectbox("选择页面", range(1, total_pages + 1), key="site_page")
                    start_idx = (page - 1) * page_size
                    end_idx = min(start_idx + page_size, len(site_names))
                    
                    current_sites = site_names[start_idx:end_idx]
                    
                    for i, site_name in enumerate(current_sites, start_idx + 1):
                        if st.button(f"{i}. {site_name}", key=f"site_btn_{i}"):
                            self._display_site_info_from_db(site_name, engineering_data)

        except Exception as e:
            st.error(f"❌ 查询工参数据失败: {e}")
            self.logger.error(f"查询工参数据失败: {e}")

    def _render_batch_export(self):
        """渲染批量导出界面"""
        st.markdown("### 📊 批量导出")

        # 从数据库获取工参数据
        try:
            # 获取所有工参数据
            engineering_data = self.db_manager.execute_query(
                "SELECT * FROM engineering_params ORDER BY phy_name, cgi"
            )
            
            if not engineering_data:
                st.warning("⚠️ 数据库中暂无工参数据，请先在总控面板导入工参数据")
                return

            # 获取所有站点名称
            site_names = list(set([row['phy_name'] for row in engineering_data if row['phy_name']]))
            site_names.sort()

            st.info(f"📊 数据库中共有 {len(engineering_data)} 条工参记录，{len(site_names)} 个站点")

            # 导出选项
            export_option = st.radio(
                "选择导出方式",
                ["导出所有站点", "导出搜索结果", "导出指定站点"],
                key="export_option"
            )

            sites_to_export = []

            if export_option == "导出所有站点":
                sites_to_export = site_names
                st.info(f"将导出所有 {len(sites_to_export)} 个站点")

            elif export_option == "导出搜索结果":
                search_query = st.text_input(
                    "输入站点名称进行搜索",
                    placeholder="支持模糊搜索...",
                    key="export_search"
                )

                if search_query:
                    sites_to_export = self._smart_search(search_query, site_names)
                    st.info(f"找到 {len(sites_to_export)} 个匹配的站点")

            elif export_option == "导出指定站点":
                selected_sites = st.multiselect(
                    "选择要导出的站点",
                    options=site_names,
                    key="export_sites"
                )
                sites_to_export = selected_sites

            if sites_to_export:
                st.write(f"准备导出 {len(sites_to_export)} 个站点")

                if st.button(
                        "📥 导出为Excel",
                        type="primary",
                        use_container_width=True):
                    try:
                        with st.spinner("正在生成Excel文件..."):
                            excel_file = self._create_excel_export_from_db(sites_to_export, engineering_data)

                            # 读取文件内容
                            with open(excel_file, 'rb') as f:
                                excel_data = f.read()

                            # 生成文件名
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename = f"工参数据_{len(sites_to_export)}个站点_{timestamp}.xlsx"

                            # 提供下载
                            st.download_button(
                                label="💾 下载Excel文件",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )

                            st.success("✅ Excel文件生成成功！")

                            # 清理临时文件
                            os.unlink(excel_file)

                    except Exception as e:
                        st.error(f"❌ 导出失败: {e}")
                        self.logger.error(f"导出工参数据失败: {e}")

        except Exception as e:
            st.error(f"❌ 查询工参数据失败: {e}")
            self.logger.error(f"查询工参数据失败: {e}")

    def _smart_search(self, query: str, site_names: list) -> list:
        """智能搜索算法"""
        query = query.strip().lower()
        if not query:
            return []

        results = []

        # 1. 精确匹配
        exact_matches = [
            name for name in site_names if str(name).lower() == query]
        results.extend(exact_matches)

        # 2. 开头匹配
        start_matches = [name for name in site_names
                         if str(name).lower().startswith(query) and name not in results]
        results.extend(start_matches)

        # 3. 包含匹配
        contains_matches = [name for name in site_names
                            if query in str(name).lower() and name not in results]
        results.extend(contains_matches)

        return results[:50]  # 最多返回50个结果

    def _display_site_info_from_db(self, site_name: str, engineering_data: list):
        """从数据库数据显示站点信息"""
        # 筛选该站点的所有工参数据
        site_data = [row for row in engineering_data if row['phy_name'] == site_name]
        
        if not site_data:
            st.error("站点不存在")
            return

        # 显示基本信息
        st.markdown("#### 📍 基本信息")
        
        # 获取站点统计信息
        total_cells = len(site_data)
        tech_types = list(set([row['zhishi'] for row in site_data if row['zhishi']]))
        frequencies = list(set([row['pinduan'] for row in site_data if row['pinduan']]))
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.write(f"**站点名称**: {site_name}")
            st.write(f"**小区总数**: {total_cells}")
        with col2:
            st.write(f"**制式**: {', '.join(tech_types) if tech_types else '未知'}")
            st.write(f"**频段**: {', '.join(frequencies) if frequencies else '未知'}")
        with col3:
            # 获取经纬度（取第一个小区的坐标）
            first_cell = site_data[0]
            st.write(f"**经度**: {first_cell['lon'] if first_cell['lon'] else '未知'}")
            st.write(f"**纬度**: {first_cell['lat'] if first_cell['lat'] else '未知'}")

        # 显示小区列表
        st.markdown("#### 📱 小区列表")
        cells_df = pd.DataFrame(site_data)
        
        # 选择要显示的列
        display_columns = ['cgi', 'celname', 'zhishi', 'pinduan', 'ant_dir',
                          'antenna_name', 'elect_tilt', 'mech_tilt', 'ant_height']
        
        # 过滤存在的列
        available_columns = [col for col in display_columns if col in cells_df.columns]
        
        if available_columns:
            st.dataframe(cells_df[available_columns], use_container_width=True, hide_index=True)
        else:
            st.dataframe(cells_df, use_container_width=True, hide_index=True)


    def _create_excel_export_from_db(self, site_names: List[str], engineering_data: List[dict]) -> str:
        """从数据库数据创建Excel导出文件"""
        # 筛选要导出的站点数据
        export_data = [row for row in engineering_data if row['phy_name'] in site_names]
        
        if not export_data:
            raise Exception("没有找到要导出的站点数据")

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "工参数据汇总"

        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = [
            "站点名称", "CGI", "小区名称", "制式", "频段", "方位角",
            "天线名称", "电下倾角", "机械下倾角", "挂高", "经度", "纬度",
            "网元状态", "机房名称", "厂家", "人力区县分公司", "站点类型", "所属规划ID"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 添加数据行
        row = 2
        for data_row in export_data:
            data = [
                data_row.get('phy_name', ''),
                data_row.get('cgi', ''),
                data_row.get('celname', ''),
                data_row.get('zhishi', ''),
                data_row.get('pinduan', ''),
                data_row.get('ant_dir', ''),
                data_row.get('antenna_name', ''),
                data_row.get('elect_tilt', ''),
                data_row.get('mech_tilt', ''),
                data_row.get('ant_height', ''),
                data_row.get('lon', ''),
                data_row.get('lat', ''),
                data_row.get('stauts_unit', ''),
                data_row.get('jifang_name', ''),
                data_row.get('manufacturer', ''),
                data_row.get('area_compy', ''),
                data_row.get('site_type', ''),
                data_row.get('pl_item', '')
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)

            row += 1

        # 调整列宽
        column_widths = [25, 20, 20, 8, 15, 8, 15, 10, 10, 8, 12, 12, 10, 15, 10, 15, 10, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动调整行高
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30

        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        return temp_file.name

    def _render_sql_export(self):
        """渲染SQL模板导出界面"""
        st.markdown("### 📋 SQL模板导出")
        st.caption("根据SQL模板查询并导出工参数据，包含beam和radius计算")
        
        try:
            # 检查数据库中是否有工参数据
            engineering_data = self.db_manager.execute_query(
                "SELECT COUNT(*) as count FROM engineering_params"
            )
            total_count = engineering_data[0]['count'] if engineering_data else 0
            
            if total_count == 0:
                st.warning("⚠️ 数据库中暂无工参数据，请先在总控面板导入工参数据")
                return
            
            st.info(f"📊 数据库中共有 {total_count} 条工参记录")
            
            # 显示SQL模板
            with st.expander("📝 查看SQL模板", expanded=False):
                st.code("""
SELECT DISTINCT
    cgi,
    celname,
    phy_name,
    antenna_name,
    stauts_unit,
    manufacturer,
    lat,
    lon,
    ant_height,
    ant_dir,
    elect_tilt,
    mech_tilt,
    site_type,
    CASE 
        WHEN site_type = '室分' THEN 359
        WHEN zhishi = '5G' AND pinduan LIKE '%700M%' THEN 40
        WHEN zhishi = '5G' AND pinduan LIKE '%2.6G%' THEN 65
        WHEN zhishi = '5G' AND pinduan LIKE '%4.9G%' THEN 70
        WHEN zhishi = '4G' AND pinduan LIKE '%FDD900%' THEN 30
        WHEN zhishi = '4G' AND pinduan LIKE '%FDD1800%' THEN 50
        WHEN zhishi = '4G' AND pinduan LIKE '%F%' THEN 45
        WHEN zhishi = '4G' AND pinduan LIKE '%D%' THEN 60
        WHEN zhishi = '4G' AND pinduan LIKE '%A%' THEN 55
        ELSE 40 
    END AS beam,
    CASE 
        WHEN site_type = '室分' THEN 30
        WHEN zhishi = '5G' AND pinduan LIKE '%700M%' THEN 50
        WHEN zhishi = '5G' AND pinduan LIKE '%2.6G%' THEN 40
        WHEN zhishi = '5G' AND pinduan LIKE '%4.9G%' THEN 30
        WHEN zhishi = '4G' AND pinduan LIKE '%FDD900%' THEN 47
        WHEN zhishi = '4G' AND pinduan LIKE '%FDD1800%' THEN 43
        WHEN zhishi = '4G' AND pinduan LIKE '%F%' THEN 39
        WHEN zhishi = '4G' AND pinduan LIKE '%D%' THEN 42
        WHEN zhishi = '4G' AND pinduan LIKE '%A%' THEN 38
        ELSE 40 
    END AS radius,
    zhishi,
    pinduan,
    pl_item
FROM engineering_params
""", language='sql')
            
            # 查询选项
            st.markdown("#### 🔍 查询选项")
            col1, col2 = st.columns(2)
            with col1:
                include_all = st.checkbox("包含所有记录", value=True)
                site_filter = st.text_input("按站点名称筛选（留空表示全部）", placeholder="输入站点名称...")
            
            with col2:
                cgi_filter = st.text_input("按CGI筛选（留空表示全部）", placeholder="输入CGI...")
                tech_filter = st.selectbox("按制式筛选", ["全部", "4G", "5G"], index=0)
            
            # 构建查询
            if st.button("🔍 执行查询", type="primary", use_container_width=True):
                with st.spinner("正在查询数据..."):
                    result_df = self._query_engineering_params_sql(
                        site_filter if site_filter else None,
                        cgi_filter if cgi_filter else None,
                        tech_filter if tech_filter != "全部" else None
                    )
                    
                    if not result_df.empty:
                        st.success(f"✅ 查询成功，共找到 {len(result_df)} 条记录")
                        
                        # 保存到session_state
                        st.session_state['sql_export_result'] = result_df
                        
                        # 显示数据预览
                        st.markdown("#### 📊 数据预览（前10条）")
                        st.dataframe(result_df.head(10), use_container_width=True, hide_index=True)
                        
                        # 导出按钮
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            excel_file_path = self._create_sql_excel_file(result_df)
                            if excel_file_path:
                                with open(excel_file_path, 'rb') as f:
                                    excel_data = f.read()
                                
                                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                filename = f"工参数据_SQL模板_{len(result_df)}条记录_{timestamp}.xlsx"
                                
                                st.download_button(
                                    label="📥 下载Excel文件",
                                    data=excel_data,
                                    file_name=filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                                
                                # 清理临时文件
                                if os.path.exists(excel_file_path):
                                    os.unlink(excel_file_path)
                        
                        with col2:
                            # 显示统计信息
                            st.caption(f"📊 总计：{len(result_df)} 条记录")
                    
                    else:
                        st.warning("⚠️ 未找到匹配的记录")
                        
        except Exception as e:
            st.error(f"❌ 查询失败: {e}")
            self.logger.error(f"SQL模板导出失败: {e}")

    def _query_engineering_params_sql(self, site_filter=None, cgi_filter=None, tech_filter=None):
        """执行SQL模板查询"""
        query = """
        SELECT DISTINCT
            cgi,
            celname,
            phy_name,
            antenna_name,
            stauts_unit,
            manufacturer,
            lat,
            lon,
            ant_height,
            ant_dir,
            elect_tilt,
            mech_tilt,
            site_type,
            CASE 
                WHEN site_type = '室分' THEN 359
                WHEN zhishi = '5G' AND pinduan LIKE '%700M%' THEN 40
                WHEN zhishi = '5G' AND pinduan LIKE '%2.6G%' THEN 65
                WHEN zhishi = '5G' AND pinduan LIKE '%4.9G%' THEN 70
                WHEN zhishi = '4G' AND pinduan LIKE '%FDD900%' THEN 30
                WHEN zhishi = '4G' AND pinduan LIKE '%FDD1800%' THEN 50
                WHEN zhishi = '4G' AND pinduan LIKE '%F%' THEN 45
                WHEN zhishi = '4G' AND pinduan LIKE '%D%' THEN 60
                WHEN zhishi = '4G' AND pinduan LIKE '%A%' THEN 55
                ELSE 40 
            END AS beam,
            CASE 
                WHEN site_type = '室分' THEN 30
                WHEN zhishi = '5G' AND pinduan LIKE '%700M%' THEN 50
                WHEN zhishi = '5G' AND pinduan LIKE '%2.6G%' THEN 40
                WHEN zhishi = '5G' AND pinduan LIKE '%4.9G%' THEN 30
                WHEN zhishi = '4G' AND pinduan LIKE '%FDD900%' THEN 47
                WHEN zhishi = '4G' AND pinduan LIKE '%FDD1800%' THEN 43
                WHEN zhishi = '4G' AND pinduan LIKE '%F%' THEN 39
                WHEN zhishi = '4G' AND pinduan LIKE '%D%' THEN 42
                WHEN zhishi = '4G' AND pinduan LIKE '%A%' THEN 38
                ELSE 40 
            END AS radius,
            zhishi,
            pinduan,
            pl_item
        FROM engineering_params
        WHERE 1=1
        """
        
        params = []
        
        if site_filter:
            query += " AND phy_name LIKE ?"
            params.append(f"%{site_filter}%")
        
        if cgi_filter:
            query += " AND cgi LIKE ?"
            params.append(f"%{cgi_filter}%")
        
        if tech_filter:
            query += " AND zhishi = ?"
            params.append(tech_filter)
        
        query += " ORDER BY phy_name, cgi"
        
        # 执行查询并转换为DataFrame
        result = self.db_manager.execute_query(query, tuple(params))
        
        if result:
            df = pd.DataFrame(result)
            # 重新排列列的顺序
            column_order = [
                'cgi', 'celname', 'phy_name', 'antenna_name', 'stauts_unit',
                'manufacturer', 'lat', 'lon', 'ant_height', 'ant_dir', 
                'elect_tilt', 'mech_tilt', 'site_type', 'beam', 'radius',
                'zhishi', 'pinduan', 'pl_item'
            ]
            # 只保留存在的列
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            # 重命名列（中文列名）
            chinese_columns = {
                'cgi': 'CGI',
                'celname': '小区名称',
                'phy_name': '物理站',
                'antenna_name': '天线名称',
                'stauts_unit': '网元状态',
                'manufacturer': '厂家',
                'lat': '纬度',
                'lon': '经度',
                'ant_height': '挂高',
                'ant_dir': '方位角',
                'elect_tilt': '电下倾角',
                'mech_tilt': '机械下倾角',
                'site_type': '站点类型',
                'beam': 'beam',
                'radius': 'radius',
                'zhishi': '制式',
                'pinduan': '频段',
                'pl_item': '所属规划ID'
            }
            df = df.rename(columns=chinese_columns)
            
            # 将所有 "nan" 替换为空值
            df = df.replace('nan', '')
            df = df.replace('NaN', '')
            df = df.replace(np.nan, '')
            # 对于数字列中的 NaN，也替换为空字符串
            df = df.where(pd.notnull(df), '')
            
            # 将空的方位角字段补充为0
            if '方位角' in df.columns:
                df['方位角'] = df['方位角'].replace('', '0')
                df['方位角'] = df['方位角'].fillna('0')
            
            return df
        else:
            return pd.DataFrame()
    
    def _create_sql_excel_file(self, df):
        """创建SQL查询结果的Excel文件（临时文件）"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "工参数据"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 添加表头
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 添加数据
            for row_idx, row_data in df.iterrows():
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx + 2, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 调整列宽
            for col in range(1, len(df.columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            return temp_file.name
            
        except Exception as e:
            self.logger.error(f"创建SQL结果Excel文件失败: {e}")
            return None
    
    def _export_sql_result_to_excel(self, df):
        """导出SQL查询结果到Excel"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "工参数据"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 添加表头
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 添加数据
            for row_idx, row_data in df.iterrows():
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx + 2, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 调整列宽
            for col in range(1, len(df.columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            # 读取文件内容
            with open(temp_file.name, 'rb') as f:
                excel_data = f.read()
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"工参数据_SQL模板_{len(df)}条记录_{timestamp}.xlsx"
            
            # 提供下载
            st.download_button(
                label="💾 下载Excel文件",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.success("✅ Excel文件生成成功！")
            
            # 清理临时文件
            os.unlink(temp_file.name)
            
        except Exception as e:
            st.error(f"❌ 导出失败: {e}")
            self.logger.error(f"导出SQL结果失败: {e}")

